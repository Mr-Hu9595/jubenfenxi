import os
import json
from typing import List, Optional

BASE_DIR = os.getcwd()
OUTPUT_PATH = os.environ.get("OUTPUT_TEMPLATE", os.path.join(BASE_DIR, "剧本评估（横维竖剧）.xlsx"))

HEADERS: List[str] = [
    "剧本名",
    "文量/集数",
    "倾向",
    "时代背景",
    "主题",
    "人物饱满度评分",
    "人物饱满度标签",
    "导演（风格建议）",
    "场景数量",
    "演员数量（核心）",
    "演员数量（群演）",
    "受众定位",
    "上线时间",
    "营销方向",
]

INPUT_JSON = os.environ.get("INPUT_JSON", os.path.join(BASE_DIR, "analysis_results.json"))

POSSIBLE_KEYS = [
    "剧本名",
    "title",
    "name",
    "script_name",
    "doc_name",
    "文件名",
    "剧本",
    "剧本名称",
]

def extract_name(record: dict) -> Optional[str]:
    for k in POSSIBLE_KEYS:
        if k in record and record[k]:
            return str(record[k])
    for k in ("file_path", "path", "filename"):
        if k in record and record[k]:
            base = os.path.basename(record[k])
            return os.path.splitext(base)[0]
    return None

def load_script_names() -> List[str]:
    names: List[str] = []
    if os.path.exists(INPUT_JSON):
        try:
            with open(INPUT_JSON, "r", encoding="utf-8") as f:
                data = json.load(f)
            records = []
            if isinstance(data, list):
                records = data
            elif isinstance(data, dict):
                for key in ("results", "items", "data"):
                    if key in data and isinstance(data[key], list):
                        records = data[key]
                        break
            for rec in records:
                n = extract_name(rec)
                if n:
                    names.append(n)
        except Exception:
            pass

    # 去重保序
    seen = set()
    ordered: List[str] = []
    for n in names:
        if n not in seen:
            ordered.append(n)
            seen.add(n)
    return ordered

def main():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except Exception as e:
        raise SystemExit(f"[ERROR] 需要 openpyxl 库来创建Excel文件: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "评估表"

    # 写入表头
    for idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="00A3A3")
        thin = Side(style="thin", color="00A3A3")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 冻结首行，便于滚动查看
    ws.freeze_panes = "A2"

    # 适配列宽（简易估算）
    widths = [14, 12, 10, 10, 12, 14, 14, 16, 12, 14, 14, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 确保目录存在
    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    # 加载剧本名并写入到A2:A行
    names = load_script_names()
    for i, name in enumerate(names, start=2):
        ws.cell(row=i, column=1, value=name)

    # 创建可筛选表格覆盖实际数据区
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        last_row = max(2, 1 + len(names))
        ref = f"A1:N{last_row}"
        table = Table(displayName="评估表", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    except Exception:
        # 回退：至少开启自动筛选
        ws.auto_filter.ref = f"A1:N{max(2, 1 + len(names))}"

    wb.save(OUTPUT_PATH)
    print(f"[OK] Created & filled: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()