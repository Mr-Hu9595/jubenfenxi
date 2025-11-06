import sys
from openpyxl import load_workbook

def main(path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    indices = [i+1 for i, v in enumerate(headers) if str(v).strip() == '测试任务标识']
    if len(indices) <= 1:
        print('[信息] 无需修复：不存在重复的“测试任务标识”列')
        return
    # 保留最后一个，删除其余（通常是倒数第二个）
    keep = max(indices)
    for idx in sorted([i for i in indices if i != keep], reverse=True):
        ws.delete_cols(idx, 1)
        print(f'[修复] 已删除多余列：索引 {idx}')
    wb.save(path)
    print('[完成] 重复列修复完成并保存')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('用法：python tools/fix_test_col.py <excel_path> <sheet_name>')
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])