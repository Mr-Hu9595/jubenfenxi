#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
为指定工作簿的目标工作表追加末列“测试任务标识”，并批量填充值与样式。

用法：
  python tools/add_test_col.py \
    --excel "/Users/mr.hu/Desktop/爆款排名/11.4 评估表.xlsx" \
    --sheet "11.4 评估输入" \
    --text "全新用户测试-11.4"

说明：
  - 自动定位当前最大列，在其后新增一列写入文本；
  - 首行写入表头“测试任务标识”；
  - 2行至数据末行统一填充值，并加粗标红；
  - 保持原冻结窗格与保护设置不变。
"""

import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', required=True)
    ap.add_argument('--sheet', required=True)
    ap.add_argument('--text', default='全新用户测试-11.4')
    args = ap.parse_args()

    wb = load_workbook(args.excel)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"[错误] 工作表不存在：{args.sheet}")
    ws = wb[args.sheet]

    # 末列后一列作为新列
    last_col_idx = ws.max_column + 1
    last_col_letter = get_column_letter(last_col_idx)

    # 表头
    ws[f"{last_col_letter}1"] = "测试任务标识"
    # 数据区（2..max_row）
    for r in range(2, ws.max_row + 1):
        c = f"{last_col_letter}{r}"
        ws[c] = args.text
        ws[c].font = Font(bold=True, color="FF0000")

    # 表头样式（红色加粗）
    ws[f"{last_col_letter}1"].font = Font(bold=True, color="FF0000")

    wb.save(args.excel)
    print(f"[完成] 已在工作表『{args.sheet}』追加列 {last_col_letter} 并写入标识：{args.text}")


if __name__ == '__main__':
    main()