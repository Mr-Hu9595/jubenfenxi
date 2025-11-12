#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
对比原模板与交付文件的《11.4 评估输入》工作表结构，输出一致性检查结果到 TXT。

生成内容包含：
  - 列名与列数、合并单元格数量、数据验证数量、条件格式数量；
  - 冻结窗格、保护状态、隐藏列/行、分组层级；
  - 追加行数（评分写入数量）；
  - 模板镜像简要（列顺序、合并范围、数据验证概要）。
  - 评估标准细则解析为结构化规则字典（从模板文件读取 A1:F32）。

用法：
  python tools/self_check_11_4.py \
    --src "/Users/mr.hu/Desktop/爆款排名/剧本评估表.xlsx" \
    --dst "/Users/mr.hu/Desktop/爆款排名/11.4 评估表.xlsx" \
    --sheet "11.4 评估输入" \
    --log "/Users/mr.hu/Desktop/爆款排名/11.4_评估日志.txt"
"""

import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def sheet_summary(ws):
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    merges = [str(rng) for rng in ws.merged_cells.ranges]
    # Data validations
    dv_list = []
    try:
        for dv in ws.data_validations.dataValidation:
            dv_list.append({
                'type': getattr(dv, 'type', None),
                'formula1': getattr(dv, 'formula1', None),
                'sqref': str(getattr(dv, 'sqref', '')),
            })
    except Exception:
        dv_list = []
    # Conditional formatting count
    cf_count = 0
    try:
        # openpyxl 3.1: ws.conditional_formatting.cf_rules is dict-like per range
        cf = ws.conditional_formatting
        if hasattr(cf, 'cf_rules'):
            cf_count = sum(len(rules) for rules in cf.cf_rules.values())
        else:
            cf_count = len(cf)
    except Exception:
        cf_count = 0
    # Freeze panes & protection
    freeze = ws.freeze_panes
    prot = {'sheet_protected': bool(getattr(ws.protection, 'sheet', False))}
    # Hidden columns/rows
    hidden_cols = sum(1 for cd in ws.column_dimensions.values() if getattr(cd, 'hidden', False))
    hidden_rows = sum(1 for rd in ws.row_dimensions.values() if getattr(rd, 'hidden', False))
    # Outline level (grouping)
    row_outline = max([getattr(rd, 'outlineLevel', 0) or 0 for rd in ws.row_dimensions.values()] or [0])
    col_outline = max([getattr(cd, 'outlineLevel', 0) or 0 for cd in ws.column_dimensions.values()] or [0])
    return {
        'rows': ws.max_row,
        'cols': ws.max_column,
        'header': header,
        'merges': merges,
        'dv_count': len(dv_list),
        'dv': dv_list,
        'cf_count': cf_count,
        'freeze': freeze,
        'protection': prot,
        'hidden_cols': hidden_cols,
        'hidden_rows': hidden_rows,
        'row_outline_level': row_outline,
        'col_outline_level': col_outline,
    }


def compare(a, b):
    diffs = []
    # 容忍末尾新增一列“测试任务标识”（任务要求第6步）
    headers_equal = a['header'] == b['header']
    cols_equal = a['cols'] == b['cols']
    if not cols_equal:
        if b['cols'] == a['cols'] + 1 and str(b['header'][-1]) == '测试任务标识':
            pass  # 允许仅新增末列
        else:
            diffs.append(f"列数不一致：src={a['cols']} dst={b['cols']}")
    if not headers_equal:
        # 允许模板列顺序前缀一致 + 末尾新增“测试任务标识”
        prefix_equal = b['header'][:a['cols']] == a['header']
        last_is_test = len(b['header']) == a['cols'] + 1 and str(b['header'][-1]) == '测试任务标识'
        if not (prefix_equal and last_is_test):
            diffs.append("列名顺序存在差异")
    if len(a['merges']) != len(b['merges']):
        diffs.append(f"合并单元格数量不同：src={len(a['merges'])} dst={len(b['merges'])}")
    if a['dv_count'] != b['dv_count']:
        diffs.append(f"数据验证数量不同：src={a['dv_count']} dst={b['dv_count']}")
    if a['cf_count'] != b['cf_count']:
        diffs.append(f"条件格式数量不同：src={a['cf_count']} dst={b['cf_count']}")
    if a['freeze'] != b['freeze']:
        diffs.append(f"冻结窗格不同：src={a['freeze']} dst={b['freeze']}")
    if a['protection'] != b['protection']:
        diffs.append(f"保护状态不同：src={a['protection']} dst={b['protection']}")
    if a['hidden_cols'] != b['hidden_cols']:
        diffs.append(f"隐藏列数量不同：src={a['hidden_cols']} dst={b['hidden_cols']}")
    if a['hidden_rows'] != b['hidden_rows']:
        diffs.append(f"隐藏行数量不同：src={a['hidden_rows']} dst={b['hidden_rows']}")
    if a['row_outline_level'] != b['row_outline_level'] or a['col_outline_level'] != b['col_outline_level']:
        diffs.append("分组层级不同")
    return diffs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--src', required=True)
    ap.add_argument('--dst', required=True)
    ap.add_argument('--sheet', default='11.4 评估输入')
    ap.add_argument('--log', required=True)
    args = ap.parse_args()

    wb_src = load_workbook(args.src, data_only=False)
    wb_dst = load_workbook(args.dst, data_only=False)
    if args.sheet not in wb_src.sheetnames:
        raise SystemExit(f"[错误] 模板中不存在工作表：{args.sheet}")
    if args.sheet not in wb_dst.sheetnames:
        raise SystemExit(f"[错误] 交付文件中不存在工作表：{args.sheet}")

    ws_src = wb_src[args.sheet]
    ws_dst = wb_dst[args.sheet]

    sum_src = sheet_summary(ws_src)
    sum_dst = sheet_summary(ws_dst)

    diffs = compare(sum_src, sum_dst)
    added_rows = sum_dst['rows'] - sum_src['rows']

    # 写 TXT 日志
    lines = []
    lines.append("==== 11.4 评估任务日志 ====")
    lines.append(f"模板工作表：{args.src}｜{args.sheet}")
    lines.append(f"交付工作表：{args.dst}｜{args.sheet}")
    lines.append("")
    lines.append("[模板镜像概要]")
    lines.append(f"列数：{sum_src['cols']} | 行数：{sum_src['rows']}")
    lines.append(f"冻结窗格：{sum_src['freeze']} | 保护：{sum_src['protection']}")
    lines.append(f"合并单元格：{len(sum_src['merges'])} | 数据验证：{sum_src['dv_count']} | 条件格式：{sum_src['cf_count']}")
    lines.append(f"隐藏列：{sum_src['hidden_cols']} | 隐藏行：{sum_src['hidden_rows']} | 分组层级：行{sum_src['row_outline_level']} 列{sum_src['col_outline_level']}")
    lines.append("列名与顺序：")
    lines.append(",".join([str(x) if x is not None else '' for x in sum_src['header']]))
    lines.append("")
    # ---------- 评估标准细则规则字典（从模板） ----------
    lines.append("[评估标准细则（结构化字典）]")
    rules = {}
    if '评估标准细则' in wb_src.sheetnames:
        ws_rules = wb_src['评估标准细则']
        for r in range(2, ws_rules.max_row + 1):
            cat = ws_rules.cell(row=r, column=1).value  # 大类
            mid = ws_rules.cell(row=r, column=2).value  # 中类
            item = ws_rules.cell(row=r, column=3).value # 小类
            val = ws_rules.cell(row=r, column=4).value  # 分值/权重
            rule = ws_rules.cell(row=r, column=5).value # 评分规则
            scope = ws_rules.cell(row=r, column=6).value# 数据口径
            cat_key = str(cat or '')
            mid_key = str(mid or '')
            rules.setdefault(cat_key, {}).setdefault(mid_key, []).append({
                '项': item,
                '值': val,
                '规则': rule,
                '口径': scope,
            })
    else:
        rules = {'提示': '模板未包含《评估标准细则》工作表'}
    # 写入到日志（简洁展开）
    for cat, mids in rules.items():
        lines.append(f"{cat}:")
        for mid, items in mids.items():
            lines.append(f"  - {mid}:")
            for it in items:
                v = it.get('值')
                lines.append(f"    * {it.get('项')}｜值={v}｜规则={it.get('规则')}｜口径={it.get('口径')}")
    lines.append("")
    lines.append("[一致性自检]")
    if not diffs:
        lines.append("一致性检查通过：结构一致")
    else:
        lines.append("一致性检查未通过：")
        lines.extend([f" - {d}" for d in diffs])
    lines.append(f"追加评分写入行数：{added_rows}")
    lines.append("")
    lines.append("[数据验证概要（模板）]")
    for i, dv in enumerate(sum_src['dv'][:20], start=1):
        lines.append(f"#{i} type={dv['type']} sqref={dv['sqref']} formula1={dv['formula1']}")
    lines.append("")
    lines.append("[数据验证概要（交付）]")
    for i, dv in enumerate(sum_dst['dv'][:20], start=1):
        lines.append(f"#{i} type={dv['type']} sqref={dv['sqref']} formula1={dv['formula1']}")

    with open(args.log, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

    print(f"[完成] 自检日志写入：{args.log}")


if __name__ == '__main__':
    main()