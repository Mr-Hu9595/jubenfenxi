#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
根据 /Users/mr.hu/Desktop/爆款排名/analysis_results_11.3.json，
为每个剧本在现有 角色表.xlsx / 场景表.xlsx 中追加新 Sheet（不覆盖旧数据）。
Sheet 名尽量使用剧本标题，若重名则添加后缀 (11.3)。
"""

import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE = "/Users/mr.hu/Desktop/爆款排名"
SRC = os.path.join(BASE, 'analysis_results_11.3.json')
ROLES_XLSX = os.path.join(BASE, '角色表.xlsx')
SCENES_XLSX = os.path.join(BASE, '场景表.xlsx')


def safe_sheet_name(name, existing):
    s = name[:31]
    if s not in existing:
        return s
    # 避免重名
    suffix = ' (11.3)'
    s2 = (name + suffix)[:31]
    if s2 not in existing:
        return s2
    i = 2
    while True:
        s3 = (name + f' (11.3-{i})')[:31]
        if s3 not in existing:
            return s3
        i += 1


def ensure_wb(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    # 删除默认空表
    wb.remove(wb.active)
    wb.save(path)
    return load_workbook(path)


def style_header(ws):
    header_fill = PatternFill("solid", fgColor="00A59D")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.freeze_panes = 'A2'


def build_roles_sheet(wb, rec):
    existing = {s.title for s in wb.sheetnames and [wb[s] for s in wb.sheetnames]}
    name = safe_sheet_name(rec['剧本名称'], existing)
    ws = wb.create_sheet(title=name)
    headers = ['角色编号','角色名称','类型','性别','年龄段','人物简介','出场场次数','台词量（估算）','演员数量需求','服化道复杂度','特殊技能/要求','关联场景ID','备注']
    ws.append(headers)
    # 角色生成
    core = rec['核心演员数']
    gender = rec['男女频']
    # 主角
    main_roles = []
    if gender in ('男','双'):
        main_roles.append(('R01','男主','男主','男','20-35','男频主角，成长/逆袭线',max(2, rec['场景数量']//4), '大', '1', '中', '打戏/普通话', 'S-核心', ''))
    if gender in ('女','双'):
        main_roles.append(('R02','女主','女主','女','20-35','女频主角，情感/事业线',max(2, rec['场景数量']//4), '大', '1', '中', '歌舞/普通话', 'S-核心', ''))
    # 反派
    main_roles.append(('R03','反派','反派','不定','25-40','主要对立面',max(2, rec['场景数量']//5), '中', '1', '中', '无', 'S-核心', ''))
    for row in main_roles:
        ws.append(list(row))
    # 配角补足
    for i in range(4, min(core, 12)+1):
        ws.append([f'R{str(i).zfill(2)}', f'配角{i-3}', '配角', '不定', '20-40', '推动情节', 1, '小', '1', '低', '无', 'S-相关', ''])
    style_header(ws)


def build_scenes_sheet(wb, rec):
    existing = {s.title for s in wb.sheetnames and [wb[s] for s in wb.sheetnames]}
    name = safe_sheet_name(rec['剧本名称'], existing)
    ws = wb.create_sheet(title=name)
    headers = ['场景编号','场景名称/描述','地点','类型（内/外）','时间（昼/夜）','体量（场次数）','群演数量','设备/技术需求','特效/后期需求','拍摄难度（0–100）','预算估算（档位）','关联角色ID','备注']
    ws.append(headers)
    # 依据题材类型生成常见场景模板
    genre = rec.get('题材类型','都市/综合')
    base_scenes = []
    if '职场' in genre or '商战' in genre:
        base_scenes = [('办公室群戏','公司','内','昼'),('会议室对峙','公司','内','昼'),('外景街拍','街道','外','昼')]
    elif '家庭' in genre:
        base_scenes = [('家庭客厅','家','内','夜'),('社区冲突','小区','外','昼'),('医院冲突','医院','内','昼')]
    elif '乡村' in genre:
        base_scenes = [('村委会','村','内','昼'),('田间劳作','田地','外','昼'),('夜谈','院坝','外','夜')]
    elif '执法' in genre:
        base_scenes = [('执法站点','路口','外','昼'),('审讯/测谎','派出所','内','夜'),('案发现场','郊外','外','夜')]
    elif '玄幻' in genre or rec['题材时代'].startswith('古代'):
        base_scenes = [('王府大殿','王府','内','昼'),('街市冲突','古城','外','昼'),('夜袭/打戏','郊外','外','夜')]
    else:
        base_scenes = [('都市室内','家','内','夜'),('公司走廊','公司','内','昼'),('街道偶遇','街道','外','昼')]

    total = max(6, rec['场景数量'])
    # 生成前三个详细 + 其余概览
    for i in range(1, min(4, len(base_scenes))+1):
        title, loc, typ, time = base_scenes[i-1]
        ws.append([f'S{str(i).zfill(2)}', title, loc, typ, time, 1, max(0, rec['群演估算']//3),
                   '轨道/斯坦尼康' if rec['技术复杂度']>40 else '基础',
                   '绿幕/特效' if rec['预算档位'] in ('高','中/高') else '常规',
                   rec['拍摄难度指数'], rec['预算档位'], 'R01,R02,R03', ''])
    for i in range(4, min(total, 12)+1):
        ws.append([f'S{str(i).zfill(2)}', '概览场景', '多地点', '不定', '不定', 1, 0,
                   '基础', '常规', rec['拍摄难度指数'], rec['预算档位'], 'R-相关', ''])

    style_header(ws)


def main():
    with open(SRC, 'r', encoding='utf-8') as f:
        records = json.load(f)
    # 保持与评估排序一致
    records.sort(key=lambda r: (
        0 if r['题材时代']=='现代' else (1 if r['题材时代'].startswith('古代') else 2),
        -r['综合优先级分'],
        -r['爆款潜力指数']
    ))

    rb = ensure_wb(ROLES_XLSX)
    sb = ensure_wb(SCENES_XLSX)
    for rec in records:
        build_roles_sheet(rb, rec)
        build_scenes_sheet(sb, rec)
    rb.save(ROLES_XLSX)
    sb.save(SCENES_XLSX)
    print(f"Appended {len(records)} sheets to 角色表.xlsx and 场景表.xlsx")


if __name__ == '__main__':
    main()