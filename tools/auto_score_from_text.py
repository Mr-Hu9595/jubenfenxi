#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
读取 /Users/mr.hu/Desktop/爆款排名/剧本评估表.xlsx 的“评估输入”工作表，
当用户提供剧本 URL（本地 txt/docx 或 http 直链）或粘贴全文时，
自动计算并填充各分项数值（子维度与得分汇总），使 Excel 公式完成总分与等级。

评分口径对齐 project_rules.md 与工作表《评估标准细则》。
"""

import os
import re
import json
from typing import Dict, Tuple
import argparse

import requests
from openpyxl import load_workbook
try:
    # 复用已有文本分析能力（时代/倾向/题材）
    from analyze_docx import detect_era, detect_gender_channel, detect_genre
except Exception:
    detect_era = None
    detect_gender_channel = None
    detect_genre = None


EXCEL_PATH = "/Users/mr.hu/Desktop/爆款排名/剧本评估表.xlsx"
SHEET_NAME = "评估输入"


# ---------- 基础读写 ----------

def _safe_read_text(path: str) -> str:
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception:
        try:
            with open(path, 'r', encoding='gb18030', errors='ignore') as f:
                return f.read()
        except Exception:
            return ""


def _read_docx_text(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return '\n'.join(p.text for p in doc.paragraphs)
    except Exception:
        # 兜底：解压 xml 文本
        import zipfile
        try:
            with zipfile.ZipFile(path) as z:
                xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
            text = re.sub(r'<[^>]+>', '', xml)
            return text
        except Exception:
            return ""


def _fetch_http_text(url: str) -> str:
    try:
        r = requests.get(url, timeout=15)
        if r.status_code == 200:
            # 简单提取文本（不处理复杂HTML）
            txt = r.text
            # 去标签
            txt = re.sub(r"<script[\s\S]*?</script>", "", txt)
            txt = re.sub(r"<style[\s\S]*?</style>", "", txt)
            txt = re.sub(r"<[^>]+>", "\n", txt)
            return re.sub(r"\n+", "\n", txt)
    except Exception:
        pass
    return ""


def read_text_from_cell(url: str, pasted: str) -> Tuple[str, str]:
    """返回 (title, text)。title 由URL/文件名或首行推断。"""
    if pasted and len(pasted.strip()) >= 50:
        text = pasted.strip()
        # 取首行作为标题候选
        first_line = text.splitlines()[0] if text.splitlines() else ""
        title = first_line[:30] if first_line else "粘贴文本"
        return title, text
    if url:
        if url.startswith('http'):
            text = _fetch_http_text(url)
            title = url.split('/')[-1]
            return title or "网页文本", text
        # 本地文件
        path = url.replace('file://', '') if url.startswith('file://') else url
        if os.path.exists(path):
            title = os.path.basename(path)
            if path.lower().endswith('.docx'):
                return title.replace('.docx',''), _read_docx_text(path)
            else:
                return title, _safe_read_text(path)
    return "", ""


# ---------- 评估计算 ----------

def count_words_k(text: str) -> float:
    return round(len(text) / 1000.0, 2)


def structure_metrics(text: str) -> Dict[str, float]:
    markers = ['引子','冲突爆发','反转','高潮','结局','序章','导火索','危机','峰值','尾声','完结']
    m_hits = sum(text.count(m) for m in markers)
    if m_hits >= 5:
        integrity = 6
    elif m_hits == 4:
        integrity = 4
    else:
        integrity = min(3, max(0, m_hits))

    triggers = ['反转','但','但是','然而','却','其实','看似','结果','谁知','出乎意料','意外']
    trig_hits = sum(text.count(t) for t in triggers)
    density_per_5k = trig_hits / max(1.0, (len(text) / 5000.0))
    if density_per_5k >= 1.0:
        twist_density = 4
    elif density_per_5k >= 0.5:
        twist_density = 3
    elif density_per_5k >= 0.25:
        twist_density = 2
    else:
        twist_density = 1 if trig_hits > 0 else 0

    foreshadow = ['伏笔','悬念','谜团','未解释线索','埋线']
    payoff = ['解释','揭示','真相','回收','说明','交代']
    f_cnt = sum(text.count(w) for w in foreshadow)
    p_cnt = sum(text.count(w) for w in payoff)
    closure = (p_cnt / f_cnt) if f_cnt > 0 else 0.0
    if closure >= 0.8:
        closure_score = 3
    elif closure >= 0.6:
        closure_score = 2
    elif f_cnt == 0 and p_cnt > 0:
        closure_score = 2
    else:
        closure_score = 1 if (f_cnt > 0 and p_cnt == 0) else 0

    # 节奏均衡度：段落字符长度的变异系数
    paras = [len(p.strip()) for p in text.split('\n') if p.strip()]
    if len(paras) >= 6:
        import statistics
        mean = statistics.mean(paras)
        stdev = statistics.pstdev(paras)
        cv = (stdev / mean) if mean > 0 else 0
        if 0.25 <= cv <= 0.8:
            rhythm = 2
        else:
            rhythm = 1
    else:
        rhythm = 1

    return {
        '起承转合完整度': integrity,
        '转折密度': twist_density,
        '悬念闭合率': closure_score,
        '节奏均衡度': rhythm,
    }


def role_metrics(text: str) -> Dict[str, float]:
    goal_words = ['目标','愿望','任务','计划']
    motive_words = ['动机','原因','缘由','初心','不得不']
    obstacle_words = ['障碍','阻碍','难题','阻止','困难','危机']
    protagonist = sum(text.count(w) for w in goal_words+motive_words+obstacle_words)
    antagonist = sum(text.count(w) for w in ['反派','敌人','对手'])
    if protagonist > 6 and antagonist > 0:
        gmo = 5
    elif protagonist > 4:
        gmo = 4
    elif protagonist > 2:
        gmo = 3
    else:
        gmo = 2 if protagonist > 0 else 1

    arc_words = ['反转','黑化','洗白','救赎','成长','挣扎','人性']
    conn_words = ['但是','然而','一方面','另一方面','看似','实则','表面','内心']
    arc_hits = sum(text.count(w) for w in arc_words)
    conn_hits = sum(text.count(w) for w in conn_words)
    if arc_hits + conn_hits >= 12:
        arc_score = 6
    elif arc_hits + conn_hits >= 7:
        arc_score = 5
    elif arc_hits + conn_hits >= 4:
        arc_score = 4
    else:
        arc_score = 3 if (arc_hits + conn_hits) >= 2 else 2

    rel_words = ['父','母','兄','姐','同事','上司','朋友','师徒','上下级','同学','同乡','婆婆','闺蜜']
    changes = ['决裂','和解','反目','结盟','分手','复合','破裂','缓和']
    unique_rel = len({w for w in rel_words if w in text})
    change_cnt = sum(text.count(w) for w in changes)
    if unique_rel >= 4 and change_cnt >= 2:
        rel_score = 3
    elif unique_rel >= 3:
        rel_score = 2
    else:
        rel_score = 1

    mem_words = ['外号','绰号','叫我','常说','口头禅','习惯','动作','独特','记忆点']
    mem_cnt = sum(text.count(w) for w in mem_words)
    mem_score = 1 if mem_cnt >= 2 else 0

    return {
        '目标-动机-障碍': gmo,
        '弧光与反差': arc_score,
        '关系网复杂度': rel_score,
        '记忆点': mem_score,
    }


def conflict_metrics(text: str) -> Dict[str, float]:
    human_conf = ['争吵','斗争','对抗','竞争','家族','敌人','反派','矛盾','冲突']
    env_conf = ['灾难','贫困','压力','体制','法律','规则','工作','公司','自然','火灾','洪水','暴雨','旱灾']
    self_conf = ['纠结','恐惧','内心','自责','自我','挣扎','心理','创伤']
    types = 0
    if any(w in text for w in human_conf):
        types += 1
    if any(w in text for w in env_conf):
        types += 1
    if any(w in text for w in self_conf):
        types += 1
    type_score = 3 if types >= 2 else (1 if types == 1 else 0)

    up_words = ['升级','加码','更','越来越','再次','逐步','层层','加深','更大','更强']
    up_cnt = sum(text.count(w) for w in up_words)
    if up_cnt >= 6:
        up_score = 3
    elif up_cnt >= 3:
        up_score = 2
    else:
        up_score = 1 if up_cnt >= 1 else 0

    risk_words = ['生命','家庭','事业','名誉','自由','金钱','财产','工作']
    cost_words = ['付出','代价','牺牲','损失','失去','破产','坐牢','赔偿']
    risk_cnt = sum(text.count(w) for w in risk_words)
    cost_cnt = sum(text.count(w) for w in cost_words)
    if risk_cnt >= 3 and cost_cnt >= 2:
        risk_score = 3
    elif risk_cnt >= 2 and cost_cnt >= 1:
        risk_score = 2
    else:
        risk_score = 1 if (risk_cnt >= 1 or cost_cnt >= 1) else 0

    strat_words = ['方案','计划','选择','策略','办法','路径','备选','B计划']
    strat_cnt = sum(text.count(w) for w in strat_words)
    strat_score = 1 if strat_cnt >= 2 else 0

    return {
        '类型覆盖': type_score,
        '升级曲线': up_score,
        '风险与代价': risk_score,
        '解决策略多样性': strat_score,
    }


def dialogue_metrics(text: str) -> Dict[str, float]:
    # 人设匹配：保守估计为 2，若存在明显风格一致的称谓/口吻则加分
    tone_consistent = ['我说','你说','他说','她说','老爷','夫人','老板','上司','兄台','小姐']
    tone_hits = sum(text.count(w) for w in tone_consistent)
    mismatch = ['忽然口音','人设崩','前后不一致']
    mismatch_hits = sum(text.count(w) for w in mismatch)
    if tone_hits >= 10 and mismatch_hits == 0:
        persona_match = 3
    elif mismatch_hits >= 1:
        persona_match = 1
    else:
        persona_match = 2

    # 信息密度：每 100 字新增有效信息点（粗略以数字/专名/地点/职业类词近似）
    info_tokens = 0
    info_tokens += len(re.findall(r"\d+", text))
    info_tokens += sum(text.count(w) for w in ['公司','项目','计划','目标','方案','事件','案','村','县','市','省','学校','医院','职位'])
    density = info_tokens / max(1, len(text)/100)
    if density >= 1.5:
        info_score = 3
    elif density >= 1.0:
        info_score = 2
    else:
        info_score = 1

    # 金句率与记忆点：每千字感叹句/引号出现次数近似衡量
    thousand = max(1, len(text)/1000)
    punchy = (text.count('！') + text.count('“')) / thousand
    if punchy >= 2.0:
        punch_score = 2
    else:
        punch_score = 1 if punchy >= 0.8 else 0

    # 生活化/专业度：是否出现专业术语或生活细节词
    domain = ['诊断','处方','诉讼','合约','融资','预算','航拍','斯坦尼康','官职','礼制']
    life = ['做饭','买菜','打车','聊天','上班','加班','学习','考试']
    prof_score = 1 if (any(w in text for w in domain) or any(w in text for w in life)) else 0

    # 价值观与合规：如存在明显不当词则 0，否则 1
    risky = ['涉黄','血腥','极端政治','恐怖主义','毒品','赌博']
    value_score = 0 if any(w in text for w in risky) else 1

    return {
        '人设匹配': persona_match,
        '信息密度': info_score,
        '金句率与记忆点': punch_score,
        '生活化/专业度': prof_score,
        '价值观与合规': value_score,
    }


def character_fullness_components(text: str) -> Dict[str, float]:
    kw_hi = ['反转','黑化','洗白','救赎']
    kw_mid = ['成长','矛盾','挣扎','人性','复杂','两面']
    connectors = ['但是','然而','一方面','另一方面','看似','实则','表面','内心']
    pairs = [('温柔','暴烈'),('理性','冲动'),('善良','狠'),('忠诚','背叛')]

    thousand = max(1.0, len(text)/1000.0)
    kw_cnt = sum(text.count(w) for w in kw_hi+kw_mid)
    conn_cnt = sum(text.count(w) for w in connectors)
    pair_cnt = sum(1 for a,b in pairs if (a in text and b in text))
    event_words = ['改变','转变','觉醒','救赎','黑化','洗白','成长']
    event_cnt = sum(text.count(w) for w in event_words)

    # 归一到 0–100（经验阈值）
    kw_weight = max(0, min(100, (kw_cnt/thousand) * 20))
    conn_weight = max(0, min(100, (conn_cnt/thousand) * 30))
    pair_weight = max(0, min(100, pair_cnt * 25))
    event_weight = max(0, min(100, event_cnt * 12))

    return {
        '人物饱满度｜关键词权重': round(kw_weight, 2),
        '人物饱满度｜连接词权重': round(conn_weight, 2),
        '人物饱满度｜反差词对': round(pair_weight, 2),
        '人物饱满度｜事件层级与弧光': round(event_weight, 2),
    }


def scene_and_cast_scores(text: str) -> Tuple[float, float]:
    # 场景数量估算（简化版），配合内外景比例词
    loc_patterns = [r'场景[一二三四五六七八九十百]+', r'第[一二三四五六七八九十百]+场']
    hits = sum(len(re.findall(p, text)) for p in loc_patterns)
    inside_kw = ['室内','屋内','内景']
    outside_kw = ['室外','街道','野外','外景','海','船','码头']
    inside = sum(text.count(w) for w in inside_kw)
    outside = sum(text.count(w) for w in outside_kw)
    total_scenes = hits if hits > 0 else max(6, min(30, inside+outside+10))
    inside_ratio = inside / max(1, inside+outside)
    # 评分口径：场景总数与内景占比（8分）
    base = 4 if total_scenes <= 12 else (6 if total_scenes <= 20 else 5)
    ratio_bonus = 2 if inside_ratio >= 0.6 else (1 if 0.45 <= inside_ratio < 0.6 else 0)
    scene_score = max(0, min(8, base + ratio_bonus))

    # 演员结构（7分）：核心角色 4–8 为佳
    role_words = ['男主','女主','反派','配角','主要角色','重要角色']
    roles = sum(text.count(w) for w in role_words)
    if roles == 0:
        roles = 6 if len(text) < 20000 else 8
    if 4 <= roles <= 8:
        cast_score = 7
    elif roles < 4:
        cast_score = 5
    else:
        cast_score = 6 if roles <= 10 else 4
    return scene_score, cast_score


def marketing_scores(title: str, text: str) -> Tuple[float, float, float]:
    # 受众定位（4）
    male = any(k in (title+text) for k in ['系统','逆袭','打脸','重生'])
    female = any(k in (title+text) for k in ['甜宠','虐恋','霸总','婚恋'])
    audience = 4 if (male and female) else (3 if (male or female) else 2)
    # 上线时间（3）：根据反转与情绪强度
    rev = sum((title+text).count(k) for k in ['反转','高潮'])
    emo = sum((title+text).count(k) for k in ['虐','甜','哭'])
    slot = 3 if rev >= 5 else (2 if emo >= 3 else 2)
    # 营销方向（3）：优先使用舆情趋势方向
    try:
        mdir = trend_marketing_direction(text, _TREND_CFG) if _TREND_CFG else None
        marketing = 3 if mdir in ('反转爽点','成长救赎','甜虐情绪') else 2
    except Exception:
        # 兜底启发式
        if any(k in text for k in ['反转','成长','救赎']):
            marketing = 3
        elif any(k in text for k in ['甜宠','虐恋']):
            marketing = 3
        elif any(k in text for k in ['知识','信息量','专业']):
            marketing = 2
        else:
            marketing = 2
    return audience, slot, marketing


def commercial_components(title: str, text: str) -> Tuple[float, float, float, float]:
    # 题材热度（2）：优先使用舆情配置映射
    try:
        hot = trend_hotness(title, text, _TREND_CFG) if _TREND_CFG else 1.0
    except Exception:
        hot = 1.0
    # 传播因子（1.5）：反转密度、情绪强度、话题词（融入趋势）
    try:
        spread = trend_spread(title, text, _TREND_CFG) if _TREND_CFG else 0.8
    except Exception:
        rev = sum(text.count(k) for k in ['反转','高潮','出乎意料','意外'])
        emo = sum(text.count(k) for k in ['虐','甜','哭','燃'])
        topic = sum(text.count(k) for k in ['话题','爆点','争议'])
        spread = max(0.5, min(1.5, rev*0.15 + emo*0.1 + topic*0.2))
    # 可生产性（1）：场景/演员/内外景比例近似
    scene_score, cast_score = scene_and_cast_scores(text)
    producible = max(0.3, min(1.0, (scene_score/8)*0.6 + (cast_score/7)*0.4))
    # 合规风险（0.5）：负面词越多风险越高，分值越低
    risky = ['涉黄','血腥','极端政治','恐怖主义','毒品','赌博']
    risk_cnt = sum(text.count(w) for w in risky)
    compliance = max(0.0, 0.5 - min(0.5, risk_cnt * 0.2))
    return hot, spread, producible, compliance


# ---------- 宣发/导演/平台推荐填充 ----------

def director_style_suggest(era: str, genre: str, text: str) -> str:
    if '刑侦' in text or '测谎' in text or '悬疑' in text:
        return '悬疑快剪'
    if era == '古代' or ('玄幻' in (genre or '')):
        if '打戏' in text or '武' in text:
            return '武侠动作'
        return '古装群像'
    if any(k in (genre or '') for k in ['职场','商战','医疗','都市']):
        return '写实都市'
    if any(k in text for k in ['系统','逆袭','打脸','反转']):
        return '商业爽剧'
    return '写实都市'


def audience_profile(gender: str, text: str, era: str) -> str:
    crowd = []
    if gender == '男':
        crowd.append('男性')
    elif gender == '女':
        crowd.append('女性')
    else:
        crowd.append('男女皆可')
    if any(k in text for k in ['农村','乡村','书记','下乡','村']):
        crowd.append('下沉')
    else:
        crowd.append('一二线')
    if any(k in text for k in ['家庭','育儿','婆婆','中年','退休']):
        crowd.append('45+')
    else:
        crowd.append('18–35')
    return '/'.join(crowd)


def online_slot(title: str, text: str) -> str:
    rev = sum((title+text).count(k) for k in ['反转','高潮'])
    emo = sum((title+text).count(k) for k in ['虐','甜','哭'])
    if rev >= 5:
        return '工作日晚间'
    if emo >= 4:
        return '周末黄金'
    return '节假日前后'


def marketing_direction(text: str) -> str:
    # 优先使用舆情配置的方向
    try:
        if _TREND_CFG:
            return trend_marketing_direction(text, _TREND_CFG)
    except Exception:
        pass
    # 兜底启发式
    if any(k in text for k in ['反转','打脸','逆袭']):
        return '反转爽点'
    if any(k in text for k in ['成长','救赎','弧光']):
        return '成长救赎'
    if any(k in text for k in ['甜宠','虐恋','霸总']):
        return '甜虐情绪'
    if any(k in text for k in ['信息量','知识','专业','术语']):
        return '知识信息量'
    if any(k in text for k in ['家国','群像','权谋']):
        return '家国叙事'
    return '反转爽点'


def platform_recommend(gender: str, text: str, era: str) -> str:
    # 简化映射：男频+反转→抖音/快手；女频+甜虐→小红书/抖音；中老年/下沉→视频号/快手。
    has_rev = any(k in text for k in ['反转','逆袭','打脸'])
    has_sweet = any(k in text for k in ['甜宠','虐恋','霸总'])
    is_rural = any(k in text for k in ['农村','乡村','书记','下乡','村'])
    audience = []
    if gender == '男' and has_rev:
        audience = ['抖音','快手']
    elif gender == '女' and has_sweet:
        audience = ['小红书','抖音']
    else:
        audience = ['抖音','快手']
    if is_rural:
        audience = ['视频号','快手']
    return '/'.join(audience)


# ---------- 主流程：填充 Excel ----------

def fill_row(ws, row_idx: int, title: str, text: str):
    if not text:
        return
    # 字数（千字）
    ws[f'E{row_idx}'] = round(len(text)/1000.0, 2)
    # 页数（粗估，1200字≈1页）
    ws[f'F{row_idx}'] = int(len(text)/1200)
    # 结构
    sm = structure_metrics(text)
    ws[f'Y{row_idx}'] = sm['起承转合完整度']
    ws[f'Z{row_idx}'] = sm['转折密度']
    ws[f'AA{row_idx}'] = sm['悬念闭合率']
    ws[f'AB{row_idx}'] = sm['节奏均衡度']
    # 角色
    rm = role_metrics(text)
    ws[f'AD{row_idx}'] = rm['目标-动机-障碍']
    ws[f'AE{row_idx}'] = rm['弧光与反差']
    ws[f'AF{row_idx}'] = rm['关系网复杂度']
    ws[f'AG{row_idx}'] = rm['记忆点']
    # 冲突
    cm = conflict_metrics(text)
    ws[f'Q{row_idx}'] = cm['类型覆盖']
    ws[f'R{row_idx}'] = cm['升级曲线']
    ws[f'S{row_idx}'] = cm['风险与代价']
    ws[f'T{row_idx}'] = cm['解决策略多样性']
    # 台词
    dm = dialogue_metrics(text)
    ws[f'AJ{row_idx}'] = dm['人设匹配']
    ws[f'U{row_idx}'] = dm['信息密度']
    ws[f'V{row_idx}'] = dm['金句率与记忆点']
    ws[f'W{row_idx}'] = dm['生活化/专业度']
    ws[f'X{row_idx}'] = dm['价值观与合规']

    # 人物饱满度四组件
    cf = character_fullness_components(text)
    ws[f'AM{row_idx}'] = cf['人物饱满度｜关键词权重']
    ws[f'AN{row_idx}'] = cf['人物饱满度｜连接词权重']
    ws[f'AO{row_idx}'] = cf['人物饱满度｜反差词对']
    ws[f'AP{row_idx}'] = cf['人物饱满度｜事件层级与弧光']

    # 制作：场景 & 演员结构
    scene_score, cast_score = scene_and_cast_scores(text)
    ws[f'L{row_idx}'] = scene_score
    ws[f'M{row_idx}'] = cast_score

    # 宣发三项分
    aud, slot, mkt = marketing_scores(title, text)
    ws[f'AT{row_idx}'] = aud
    ws[f'AU{row_idx}'] = slot
    ws[f'AV{row_idx}'] = mkt

    # 商业价值四组件
    hot, spread, prod, comp = commercial_components(title, text)
    ws[f'AX{row_idx}'] = hot
    ws[f'AY{row_idx}'] = spread
    ws[f'AZ{row_idx}'] = prod
    ws[f'BA{row_idx}'] = comp

    # 复用检测：倾向/时代/主题
    try:
        if detect_gender_channel:
            gender = detect_gender_channel(text, title)
            ws[f'H{row_idx}'] = {'男':'男频','女':'女频'}.get(gender, '双频')
        if detect_era:
            era = detect_era(text, title)
            # 统一到规则口径
            era_map = {'现代':'现代','古代':'古代','玄幻/奇幻':'科幻/奇幻'}
            ws[f'I{row_idx}'] = era_map.get(era, '其他')
        if detect_genre:
            era_val = ws[f'I{row_idx}'].value or '现代'
            genre = detect_genre(text, title, era_val)
            ws[f'J{row_idx}'] = genre
    except Exception:
        pass

    # 导演风格建议/受众定位/上线时间/营销方向/平台推荐
    era_val = (ws[f'I{row_idx}'].value or '现代')
    genre_val = (ws[f'J{row_idx}'].value or '')
    gender_val = ws[f'H{row_idx}'].value or '双频'
    ws[f'K{row_idx}'] = director_style_suggest(era_val, genre_val, text)
    ws[f'N{row_idx}'] = audience_profile('男' if '男' in gender_val else ('女' if '女' in gender_val else '双'), text, era_val)
    ws[f'O{row_idx}'] = online_slot(title, text)
    ws[f'P{row_idx}'] = marketing_direction(text)
    ws[f'BH{row_idx}'] = platform_recommend('男' if '男' in gender_val else ('女' if '女' in gender_val else '双'), text, era_val)


def main():
    parser = argparse.ArgumentParser(description="批量解析文本并填充评估输入工作表")
    parser.add_argument("--sheet", default=SHEET_NAME, help="目标工作表名称，默认“评估输入”")
    parser.add_argument("--excel", default=EXCEL_PATH, help="Excel 文件路径")
    args = parser.parse_args()

    wb = load_workbook(args.excel)
    if args.sheet not in wb.sheetnames:
        raise RuntimeError(f"工作表不存在：{args.sheet}")
    ws = wb[args.sheet]
    # 扫描 2..max 行
    max_row = ws.max_row
    for r in range(2, max_row+1):
        url = ws[f'C{r}'].value or ''
        pasted = ws[f'D{r}'].value or ''
        title, text = read_text_from_cell(url, pasted)
        if not text:
            continue
        # 写入剧本名（若为空）
        if not ws[f'B{r}'].value:
            ws[f'B{r}'] = title[:50]
        fill_row(ws, r, title, text)

    wb.save(args.excel)
    print(f"已完成自动填充：{args.excel} -> 工作表：{args.sheet}，处理行数 {max_row-1}")


if __name__ == '__main__':
    main()
# 舆情模块（可选导入）
try:
    from trending import load_trending_config, trend_hotness, trend_spread, trend_marketing_direction
    _TREND_CFG = load_trending_config()
except Exception:
    _TREND_CFG = None