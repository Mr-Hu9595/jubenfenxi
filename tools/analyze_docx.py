#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
递归解析指定目录下的 docx 文件，抽取关键信息并计算评分与综合排序；
同时可生成 Excel 与分析溯源 JSON（路径可通过环境变量覆盖）。

环境变量：
- DOCX_DIR：待解析的 docx 目录（默认 samples/）。
- OUT_JSON：分析结果输出 JSON（默认工作目录 analysis_results.json）。
- OUT_XLSX：结果 Excel 输出路径（默认工作目录 剧本评估表.xlsx）。

字段与算法参考 tools/analyze_scripts.py，并增强题材类型识别与 docx 解析。
"""

import os
import re
import json
from collections import Counter

BASE_DIR = os.getcwd()
DOCX_DIR = os.environ.get("DOCX_DIR", os.path.join(BASE_DIR, "samples"))
OUT_JSON = os.environ.get("OUT_JSON", os.path.join(BASE_DIR, "analysis_results.json"))
OUT_XLSX = os.environ.get("OUT_XLSX", os.path.join(BASE_DIR, "剧本评估表.xlsx"))


def list_docx_files(root):
    files = []
    for dirpath, dirnames, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower().endswith('.docx'):
                files.append(os.path.join(dirpath, fn))
    return sorted(files)


def read_docx_text(path):
    """优先使用 python-docx 读取；若不可用则尝试压缩包解析兜底。"""
    try:
        from docx import Document
        doc = Document(path)
        return '\n'.join(p.text for p in doc.paragraphs)
    except Exception:
        try:
            import zipfile
            with zipfile.ZipFile(path) as z:
                xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
            # 去掉 xml 标签，保留文字
            text = re.sub(r'<[^>]+>', '', xml)
            return text
        except Exception:
            return ""


def detect_era(text, title):
    ancient_kw_title = ['王府','侯府','皇帝','太子','皇位','朝臣','宫','锦衣','科举','状元','育儿师','道爷']
    fantasy_kw_title = ['仙','修仙','仙帝','魔族','女帝','帝','妖','神','黄毛','反派']
    modern_kw_title = ['大学','直播','抖音','快手','视频号','系统','公司','上司','交警','书记','豪门','外卖','村','毕业','网红','主播','拼夕夕','海蛇','船','测谎']

    if any(k in title for k in ancient_kw_title):
        return '古代'
    if any(k in title for k in fantasy_kw_title):
        return '玄幻/奇幻'
    if any(k in title for k in modern_kw_title):
        return '现代'

    ancient_keywords = ['古代','王府','侯府','皇帝','太子','皇位','朝臣','宫','锦衣','科举','状元','民间']
    fantasy_keywords = ['仙','修仙','仙帝','魔族','女帝','帝尊','玄幻','妖','神']
    t = text
    if any(k in t for k in ancient_keywords):
        return '古代'
    if any(k in t for k in fantasy_keywords):
        return '玄幻/奇幻'
    return '现代'


def detect_gender_channel(text, title):
    female_hits = [
        '甜宠','虐恋','言情','豪门女主','女主','男友','恋','上司','育儿师','报仇','霸总','婚恋','婆婆','闺蜜','保镖'
    ]
    male_hits = [
        '系统','重生','开局','打脸','逆袭','升级','皇帝','科举','状元','摆烂','直播','攻略','捞','捕鱼','卧底','逼婚','皇位','老祖','船','海'
    ]
    ft = title + " " + text
    f_score = sum(ft.count(k) for k in female_hits)
    m_score = sum(ft.count(k) for k in male_hits)
    if f_score > m_score * 1.2:
        return '女'
    if m_score > f_score * 1.2:
        return '男'
    return '双'


def detect_original(text):
    return '改编' if ('改编' in text or ('版权' in text and '改' in text)) else '原创'


def detect_genre(text, title, era):
    ft = (title + '\n' + text)
    # 先判断大类
    if '修仙' in ft or '仙帝' in ft or '魔族' in ft or '女帝' in ft or '玄幻' in ft:
        return '玄幻修仙'
    if era == '古代' and any(k in ft for k in ['皇帝','王府','朝堂','科举','侯府']):
        return '古代权谋/科举'
    # 现代细分
    if any(k in ft for k in ['公司','上司','领导','加班','投资','资本家','老板']):
        return '职场/商战'
    if any(k in ft for k in ['老婆','姐姐','离婚','家庭','后妈','育儿师','婆婆']):
        return '家庭/婚恋'
    if any(k in ft for k in ['农村','乡村','村','书记','GDP']):
        return '乡村/三农'
    if any(k in ft for k in ['学校','大学','毕业','校花','状元（现代）']):
        return '校园/成长'
    if any(k in ft for k in ['交警','测谎','卧底','刑侦','执法']):
        return '执法/社会话题'
    if any(k in ft for k in ['医生','行医','医院','神医']):
        return '医疗/都市'
    if any(k in ft for k in ['直播','视频号','网红','整活']):
        return '娱乐/直播'
    if any(k in ft for k in ['系统','首富','神豪','亿亿神豪']):
        return '都市/系统神豪'
    # 兜底
    return '都市/综合'


def count_core_roles(text):
    keys = ['男主','女主','反派','配角','重要角色','主要角色']
    c = sum(text.count(k) for k in keys)
    length = len(text)
    if c == 0:
        if length < 8000:
            return 4
        elif length < 15000:
            return 6
        elif length < 30000:
            return 8
        else:
            return 10
    return max(4, min(12, c))


def count_scenes(text):
    patterns = [
        r'场景[一二三四五六七八九十百]+', r'第[一二三四五六七八九十百]+场', r'内景', r'外景', r'昼', r'夜'
    ]
    hits = 0
    for p in patterns:
        hits += len(re.findall(p, text))
    if hits == 0:
        loc_words = ['家','公司','医院','学校','村','街道','王府','皇宫','科举考场','牢房','郊外','码头','船']
        loc_hits = sum(text.count(k) for k in loc_words)
        hits = max(6, min(20, loc_hits))
    if hits < 6:
        return 6
    if hits > 30:
        return 30
    return hits


def budget_bracket(era, text):
    high_words = ['皇帝','王府','朝臣','宫','航拍','大型群演','打戏','水戏','海','船','特效','绿幕']
    medium_words = ['夜戏','外景','车戏','办公室群戏','学校群戏']
    if era.startswith('古代'):
        if any(w in text for w in high_words):
            return '高'
        return '中'
    if '玄幻' in era:
        return '高' if any(w in text for w in high_words) else '中'
    if any(w in text for w in high_words):
        return '中/高'
    if any(w in text for w in medium_words):
        return '中'
    return '低'


def tech_complexity_score(text):
    tech_words = ['航拍','斯坦尼康','轨道','绿幕','特效','打戏','水下','船','海','爆破','枪战']
    hits = sum(text.count(w) for w in tech_words)
    return max(0, min(100, hits * 12))


def novelty_score(title, text):
    novel_words = ['分身','摆烂系统','拼夕夕','海蛇','饥荒年','育儿师','女帝逼婚','视频号','直播逼我整活','捞船装备','黄毛','民国','政务','测谎']
    common_words = ['系统','重生','科举','皇帝','状元','逆袭','打脸','豪门','保镖','外卖']
    n = sum((title+text).count(w) for w in novel_words)
    c = sum((title+text).count(w) for w in common_words)
    base = 6
    if n >= 3:
        base = 8
    elif n == 2:
        base = 7
    elif n == 1:
        base = 6.5
    else:
        base = 5.5
    base -= min(2.5, c * 0.15)
    base = max(3.5, min(9.5, base))
    return round(base * 10)  # 0-100


def market_potential_score(era, gender, title, text):
    boosts = 0
    if era == '现代':
        boosts += 10
    if gender == '男' and any(k in (title+text) for k in ['系统','重生','逆袭','打脸','直播']):
        boosts += 15
    if gender == '女' and any(k in (title+text) for k in ['甜宠','虐恋','霸总','婚恋','男友']):
        boosts += 15
    if '乡村' in text or '书记' in text or '清北' in text or '农村' in text:
        boosts += 8
    if '科举' in text or '皇帝' in text:
        boosts += 5
    base = 60 + boosts
    base = max(35, min(95, base))
    return base


def difficulty_index(core_roles, scenes, budget, tech_score):
    role_s = max(10, min(100, int(core_roles / 12 * 100)))
    scene_s = max(10, min(100, int(scenes / 30 * 100)))
    budget_s_map = {'低': 30, '中': 60, '中/高': 75, '高': 90}
    budget_s = budget_s_map.get(budget, 60)
    tech_s = tech_score
    geo_s = int(scene_s * 0.8)
    idx = role_s*0.25 + scene_s*0.25 + budget_s*0.25 + tech_s*0.15 + geo_s*0.10
    return round(idx)


def difficulty_star(idx):
    if idx < 35:
        return '★☆'
    if idx < 50:
        return '★★'
    if idx < 65:
        return '★★★'
    if idx < 80:
        return '★★★★'
    return '★★★★★'


def budget_control_score(budget):
    return {'低': 90, '中': 65, '中/高': 55, '高': 40}.get(budget, 60)


def overall_priority(potential, difficulty, budget_control, novelty, era):
    score = potential*0.35 + (100 - difficulty)*0.25 + budget_control*0.20 + novelty*0.15 + (5 if era == '现代' else 0)
    return round(score)


def character_depth_score(title, text):
    """估算人物饱满度（0–100，多面性指数）
    统一复用 auto_score_from_text 的四组件计算，并按 40/30/20/10 加权。
    """
    try:
        # 延迟导入以避免循环依赖
        from auto_score_from_text import character_fullness_components
        comp = character_fullness_components(title + "\n" + text)
        score = (
            (comp.get('人物饱满度｜关键词权重', 0) * 0.4)
            + (comp.get('人物饱满度｜连接词权重', 0) * 0.3)
            + (comp.get('人物饱满度｜反差词对', 0) * 0.2)
            + (comp.get('人物饱满度｜事件层级与弧光', 0) * 0.1)
        )
        return round(max(0.0, min(100.0, score)), 2)
    except Exception:
        # 兜底：若导入失败，返回中位值以不中断流程
        return 50.0


def character_depth_label(score):
    if score >= 67:
        return '高'
    if score >= 34:
        return '中'
    return '低'


def analyze_one(path):
    title = os.path.basename(path).replace('.docx','')
    text = read_docx_text(path)
    era = detect_era(text, title)
    gender = detect_gender_channel(text, title)
    original = detect_original(text)
    genre = detect_genre(text, title, era)
    core_roles = count_core_roles(text)
    scenes = count_scenes(text)
    budget = budget_bracket(era, text)
    tech_s = tech_complexity_score(text)
    diff_idx = difficulty_index(core_roles, scenes, budget, tech_s)
    diff_star = difficulty_star(diff_idx)
    novelty = novelty_score(title, text)
    # 人物饱满度与平台舆情加权（红果/抖音/快手/小红书趋势：反转+成长更受欢迎）
    char_depth = character_depth_score(title, text)
    char_label = character_depth_label(char_depth)
    potential = market_potential_score(era, gender, title, text)
    if char_depth >= 67:
        potential = min(100, potential + 8)
    elif char_depth >= 34:
        potential = min(100, potential + 4)
    budget_ctrl = budget_control_score(budget)
    overall = overall_priority(potential, diff_idx, budget_ctrl, novelty, era)

    return {
        '剧本名称': title,
        '男女频': gender,
        '是否原创': original,
        '题材时代': era,
        '题材类型': genre,
        '核心演员数': core_roles,
        '群演估算': max(0, int(core_roles*0.6)),
        '场景数量': scenes,
        '预算档位': budget,
        '技术复杂度': tech_s,
        '拍摄难度指数': diff_idx,
        '拍摄难易度': diff_star,
        '题材新颖度': novelty,            # 0-100，写表时换算 1-10
        '爆款潜力指数': potential,        # 0-100，写表时换算 1-10
        '人物饱满度': char_depth,
        '人物饱满度标签': char_label,
        '预算可控性': budget_ctrl,
        '综合优先级分': overall,
    }


def _col_letter(idx):
    import string
    letters = list(string.ascii_uppercase)
    if idx <= 26:
        return letters[idx-1]
    # 支持到两位字母即可
    first = letters[(idx-1)//26 - 1]
    second = letters[(idx-1)%26]
    return first + second


def write_excel(records):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = '综合排名'

    headers = [
        '排名','剧本名称','男女频','是否原创','题材类型',
        '演员数量（核心/群演）','场景数量','拍摄难易度','预算成本（低/中/高）',
        '题材新颖度（10分制）','人物饱满度（高/中/低）','爆款潜力指数（10分制）','综合优先级分','备注摘要'
    ]
    ws.append(headers)

    for i, r in enumerate(records, 1):
        novelty10 = max(1, min(10, round(r['题材新颖度']/10)))
        potential10 = max(1, min(10, round(r['爆款潜力指数']/10)))
        actors_cell = f"{r['核心演员数']}人/{r['群演估算']}人"
        char_label = r.get('人物饱满度标签') or character_depth_label(r.get('人物饱满度', 50))
        remark = (
            f"{r['题材时代']}、{r['题材类型']}；男女频：{r['男女频']}；成本：{r['预算档位']}；"
            f"拍摄指数{r['拍摄难度指数']}（{r['拍摄难易度']}）；平台推荐："
            + ("抖音/快手" if r['题材时代']=='现代' else "视频号/多平台")
        )
        row = [
            i, r['剧本名称'], r['男女频'], r['是否原创'], r['题材类型'],
            actors_cell, r['场景数量'], r['拍摄难易度'], r['预算档位'],
            novelty10, char_label, potential10, r['综合优先级分'], remark
        ]
        ws.append(row)

    # 列宽
    col_widths = [6, 36, 8, 10, 16, 18, 10, 14, 16, 16, 12, 18, 16, 60]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[_col_letter(idx)].width = w

    # 表头样式（青绿色底、白色粗体、居中）
    header_fill = PatternFill("solid", fgColor="00A59D")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 数据居中（适度）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = center

    # 冻结首行 & 启用筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{_col_letter(len(headers))}{ws.max_row}"

    wb.save(OUT_XLSX)
    return OUT_XLSX, ws.max_row, len(headers)


def main():
    paths = list_docx_files(DOCX_DIR)
    records = []
    for p in paths:
        rec = analyze_one(p)
        records.append(rec)
    # 排序：现代剧优先 -> 综合优先级 -> 爆款潜力
    records.sort(key=lambda r: (
        0 if r['题材时代']=='现代' else (1 if r['题材时代'].startswith('古代') else 2),
        -r['综合优先级分'],
        -r['爆款潜力指数']
    ))
    # 加排名
    for i, r in enumerate(records, 1):
        r['排名'] = i

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    xlsx_path, rows, cols = write_excel(records)
    print(f"Wrote {len(records)} records to {OUT_JSON}")
    print(f"Excel: {xlsx_path}, rows={rows}, cols={cols}")


if __name__ == '__main__':
    main()