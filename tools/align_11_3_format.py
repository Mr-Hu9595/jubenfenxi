#!/usr/bin/env python3
"""
Align the "11.3 项目评估.xlsx" to the reference sheet1 format:
- Create a new sheet named "综合排名（参考版）" with headers strictly matching the reference.
- Load data from analysis_results_11.3.json, map fields, and compute display values
  (star rating for difficulty, tier labels, novelty category, 10-point potential index).
- Freeze the first row and apply basic column widths. Header styling will be applied
  externally via tooling to keep consistency.

Assumptions:
- analysis_results_11.3.json contains per-script metrics similar to previous pipeline.
- Keys may vary; this script includes robust getters and fallbacks.
"""

import json
import math
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path("/Users/mr.hu/Desktop/爆款排名")
JSON_PATH = ROOT / "analysis_results_11.3.json"
XLSX_PATH = ROOT / "11.3 项目评估.xlsx"


REF_SHEET_NAME = "综合排名（参考版）"


def _get(obj, *keys, default=None):
    for k in keys:
        if isinstance(k, (list, tuple)):
            for kk in k:
                if kk in obj:
                    return obj.get(kk)
        else:
            if k in obj:
                return obj.get(k)
    return default


def difficulty_to_stars(difficulty_index: float) -> str:
    if difficulty_index is None:
        return "★★★"
    try:
        x = float(difficulty_index)
    except Exception:
        return "★★★"
    # Map 0-100 to 1-5 stars
    if x <= 0:
        n = 1
    else:
        n = max(1, min(5, math.ceil(x / 20)))
    return "★" * n


def novelty_to_label(novelty_score: float) -> str:
    if novelty_score is None:
        return "中"
    try:
        x = float(novelty_score)
    except Exception:
        return "中"
    if x >= 67:
        return "高"
    elif x >= 34:
        return "中"
    else:
        return "低"


def potential_to_10(hit_potential_score: float) -> float:
    if hit_potential_score is None:
        return 5.0
    try:
        x = float(hit_potential_score)
    except Exception:
        return 5.0
    # convert 0-100 to 0-10
    return round(x / 10.0, 1)


def character_depth_to_label(depth_score, depth_label=None) -> str:
    if depth_label:
        s = str(depth_label)
        if any(k in s for k in ["高", "中", "低"]):
            return s
    if depth_score is None:
        return "中"
    try:
        x = float(depth_score)
    except Exception:
        return "中"
    if x >= 67:
        return "高"
    elif x >= 34:
        return "中"
    else:
        return "低"


def audience_to_label(audience: str) -> str:
    if not audience:
        return "双频"
    s = str(audience).lower()
    if "男" in s or "male" in s:
        return "男频"
    if "女" in s or "female" in s:
        return "女频"
    if "双" in s or "both" in s or "混合" in s:
        return "双频"
    return "双频"


def original_to_label(val) -> str:
    if isinstance(val, bool):
        return "原创" if val else "改编"
    if val is None:
        return "原创"
    s = str(val)
    if "原创" in s:
        return "原创"
    if "改编" in s or "adapt" in s.lower():
        return "改编"
    return "原创"


def budget_to_tier(val) -> str:
    if val is None:
        return "中"
    s = str(val)
    if any(k in s for k in ["高", "High", "high"]):
        return "高"
    if any(k in s for k in ["低", "Low", "low"]):
        return "低"
    if any(k in s for k in ["中", "Medium", "medium"]):
        return "中"
    # If numeric range present, infer tiers by value
    try:
        x = float(val)
        if x >= 0.66:
            return "高"
        elif x >= 0.33:
            return "中"
        else:
            return "低"
    except Exception:
        return "中"


def make_summary(item: dict) -> str:
    era = _get(item, "时代", "era", default="现代")
    genre = _get(item, "题材类型", "genre", default="都市/职场")
    budget = budget_to_tier(_get(item, "预算档位", "budget_tier", "budget", default="中"))
    diff_idx = _get(item, "拍摄难度指数", "difficulty_index", default=50)
    potential = potential_to_10(_get(item, "爆款潜力指数", "hit_potential_index", "hitPotential", default=60))
    # Brief rationale sentence
    return f"{era}+{genre}；难度{difficulty_to_stars(diff_idx)}，预算{budget}，爆款潜力{potential}分；建议关注抖音/视频号。"


def main():
    if not JSON_PATH.exists():
        raise FileNotFoundError(f"Missing {JSON_PATH}")
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing {XLSX_PATH}")

    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    # data could be a list or dict with 'items'
    if isinstance(data, dict) and "items" in data:
        items = data["items"]
    else:
        items = data

    # Sort by composite priority score when available, otherwise by hit potential
    def sort_key(it: dict):
        return (
            _get(it, "综合优先级分", "composite_priority_score", "priority_score", default=0),
            _get(it, "爆款潜力指数", "hit_potential_index", "hitPotential", default=0),
        )

    try:
        items_sorted = sorted(items, key=sort_key, reverse=True)
    except Exception:
        items_sorted = items

    wb: Workbook = load_workbook(XLSX_PATH)
    # Remove existing ref sheet if any to avoid duplicates
    if REF_SHEET_NAME in wb.sheetnames:
        ws_existing = wb[REF_SHEET_NAME]
        wb.remove(ws_existing)

    ws: Worksheet = wb.create_sheet(REF_SHEET_NAME)

    # Headers strictly follow the reference screenshot
    headers = [
        "排名",
        "剧本名称",
        "男女频",
        "是否原创",
        "题材类型",
        "演员数量（核心/群演）",
        "场景数量",
        "拍摄难易度（★★★★★）",
        "预算成本（低/中/高）",
        "题材新颖度（高/中/低）",
        "人物饱满度（高/中/低）",
        "爆款潜力指数（10分制）",
        "总结/备注",
    ]

    ws.append(headers)

    for idx, item in enumerate(items_sorted, start=1):
        name = _get(item, "剧本名", "剧本名称", "script_name", "title", default=f"剧本{idx}")
        audience = audience_to_label(_get(item, "男女频", "audience", "gender_frequency", default="双频"))
        original_label = original_to_label(_get(item, "是否原创", "original", default=True))
        genre = _get(item, "题材类型", "genre", default="都市/职场")

        # Actor counts: prefer explicit fields; fall back to estimates
        actors_core = _get(item, "核心演员数量", "actors_core", "actors_main", "演员数量", "actors_estimate", default=6)
        extras = _get(item, "群演数量", "extras_estimate", "extras", default=5)
        actors_display = f"{int(actors_core)}人<{int(extras)}人"

        scenes = _get(item, "场景数量", "scenes_estimate", "scene_count", default=8)
        diff_idx = _get(item, "拍摄难度指数", "difficulty_index", default=50)
        diff_stars = difficulty_to_stars(diff_idx)
        budget_tier = budget_to_tier(_get(item, "预算档位", "budget_tier", "budget", default="中"))
        novelty_label = novelty_to_label(_get(item, "题材新颖度", "novelty_score", default=60))
        char_label = character_depth_to_label(
            _get(item, "人物饱满度", "character_depth", default=60),
            _get(item, "人物饱满度标签", "character_depth_label", default=None)
        )
        potential_10 = potential_to_10(_get(item, "爆款潜力指数", "hit_potential_index", "hitPotential", default=60))

        summary = _get(item, "总结", "备注", "recommendation", default=None)
        if not summary:
            summary = make_summary(item)

        row = [
            idx,
            name,
            audience,
            original_label,
            genre,
            actors_display,
            scenes,
            diff_stars,
            budget_tier,
            novelty_label,
            char_label,
            potential_10,
            summary,
        ]
        ws.append(row)

    # Freeze the header row and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    # Basic column widths for readability
    widths = [6, 28, 10, 12, 18, 18, 10, 18, 16, 18, 14, 20, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(XLSX_PATH)
    print(f"Aligned sheet '{REF_SHEET_NAME}' written to: {XLSX_PATH}")


if __name__ == "__main__":
    main()