#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
通用 CLI：批量导入剧本文本（txt/docx/pdf），自动填充剧本评估 Excel。

功能概述：
- 递归扫描指定目录或处理单文件，支持 .txt/.docx/.pdf。
- 读取文本后调用现有评估逻辑（fill_row）写入指定工作表。
- 自动补齐关键联动公式（建议集数、分项汇总、人物饱满度、商业价值指数、总分、评分等级）。
- 若目标工作表不存在，将复制“评估输入”的表头创建新表。

使用示例：
python tools/universal_cli.py --input ./scripts --excel \
  "/Users/mr.hu/Desktop/爆款排名/剧本评估表.xlsx" --sheet "通用导入"

注意：
- PDF 解析优先使用 pdfminer.six；如不可用或解析失败，按规则提示提供可解析文本版本。
"""

import os
import sys
import re
import argparse
from typing import List, Tuple, Dict

# 可选加载 jieba 分词（用于摘要优化）。未安装时自动回退。
try:
    import jieba  # type: ignore
    JIEBA_AVAILABLE = True
except Exception:
    JIEBA_AVAILABLE = False

# 可选加载 jieba.analyse 与词性分词（用于更精细权重）
if JIEBA_AVAILABLE:
    try:
        from jieba import analyse as jieba_analyse  # type: ignore
        import jieba.posseg as pseg  # type: ignore
        JIEBA_ANALYSE_AVAILABLE = True
    except Exception:
        JIEBA_ANALYSE_AVAILABLE = False
else:
    JIEBA_ANALYSE_AVAILABLE = False

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 确保可导入同目录下的 auto_score_from_text
HERE = os.path.dirname(__file__)
if HERE not in sys.path:
    sys.path.insert(0, HERE)

try:
    from auto_score_from_text import fill_row  # 复用已有评分逻辑
except Exception as e:
    print("[错误] 无法导入 auto_score_from_text.fill_row：", e)
    print("请确保在 tools/ 目录内存在 auto_score_from_text.py 并可被导入。")
    sys.exit(1)

# 可选加载题材/倾向识别（用于摘要风格细分）
try:
    from analyze_docx import detect_era, detect_gender_channel, detect_genre  # type: ignore
except Exception:
    detect_era = None
    detect_gender_channel = None
    detect_genre = None


def read_txt(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="gbk", errors="ignore") as f:
            return f.read()
    except Exception:
        return ""


def read_docx(path: str) -> str:
    # 优先使用 python-docx；失败时尝试解压 xml
    try:
        import docx  # type: ignore
        doc = docx.Document(path)
        # 读取文档段落文本
        body_text = "\n".join([p.text for p in doc.paragraphs])
        # 读取文档 metadata 标题，供 extract_title 使用（避免回退数字文件名）
        meta_title = ""
        try:
            meta_title = (getattr(doc, 'core_properties', None).title or '').strip()
        except Exception:
            meta_title = ""
        # 若存在有效 metadata 标题，将其作为显式标签注入文本前缀
        if meta_title and len(meta_title) >= 2:
            prefix = f"标题：{meta_title}\n"
            return prefix + body_text
        return body_text
    except Exception:
        # 退回使用 zip 解包 word/document.xml
        import zipfile
        try:
            with zipfile.ZipFile(path) as z:
                xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
            # 简单去标签
            return (
                xml.replace("<w:p>", "\n")
                .replace("</w:p>", "")
                .replace("<w:t>", "")
                .replace("</w:t>", "")
            )
        except Exception:
            return ""


def read_pdf(path: str) -> Tuple[str, str]:
    """
    返回 (文本, 提示)。若无法解析，文本为空并提供提示信息。
    """
    # 优先 pdfminer.six
    try:
        from pdfminer.high_level import extract_text  # type: ignore
        text = extract_text(path) or ""
        if text.strip():
            return text, ""
    except Exception:
        pass

    # 备选 PyMuPDF
    try:
        import fitz  # type: ignore
        doc = fitz.open(path)
        texts = []
        for page in doc:
            texts.append(page.get_text())
        text = "\n".join(texts)
        if text.strip():
            return text, ""
    except Exception:
        pass

    # 备选 pdfplumber
    try:
        import pdfplumber  # type: ignore
        texts = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                texts.append(page.extract_text() or "")
        text = "\n".join(texts)
        if text.strip():
            return text, ""
    except Exception:
        pass

    return "", "PDF 解析失败：请提供可解析的 txt/docx 文本版本。"


def extract_title(text: str, fallback_name: str = "") -> str:
    """从文本中智能提取剧本标题，优先解析《书名号》或显式标签。

    规则（按优先级）：
    - 书名号：匹配首 3000 字中的《...》内容，长度 2–40。
    - 显式标签：剧名/片名/标题/作品名/书名/剧本名 后的值（2–40）。
    - 首行候选：前 20 行中最像标题的一行（含中文、长度 3–40，排除作者/编剧/目录/章节等）。
    - 兜底：返回 fallback_name。
    """
    try:
        t = (text or "").strip()
        if not t:
            return fallback_name
        head = t[:3000]
        # 1) 书名号
        m = re.search(r"《([\u4e00-\u9fa5A-Za-z0-9\-\s·—_]{2,40})》", head)
        if m:
            title = m.group(1).strip()
            return title
        # 2) 显式标签
        tag_pat = re.compile(
            r"(剧名|片名|标题|作品名|书名|剧本名)[：:]\s*([\u4e00-\u9fa5A-Za-z0-9\-·—_\s]{2,40})"
        )
        m2 = tag_pat.search(head)
        if m2:
            title = m2.group(2).strip()
            # 清理分隔符与尾随说明
            title = re.split(r"[\|/\\]", title)[0].strip()
            return title
        # 3) 首行候选
        lines = [ln.strip() for ln in t.splitlines() if ln.strip()][:20]
        def is_candidate(ln: str) -> bool:
            if len(ln) < 3 or len(ln) > 40:
                return False
            # 排除常见非标题行
            bad = ["作者", "编剧", "出品", "出品人", "日期", "目录", "版权", "序言", "前言",
                    "第一章", "第二章", "第", "章", "集", "分集", "大纲", "目录"]
            if any(b in ln for b in bad):
                return False
            # 必须包含中文或书名号
            if not re.search(r"[\u4e00-\u9fa5]", ln):
                return False
            # 排除纯数字/编号
            if re.fullmatch(r"[0-9\-_.\s]+", ln):
                return False
            return True
        for ln in lines:
            if is_candidate(ln):
                # 去除常见尾缀
                ln = re.sub(r"[\s·—\-]+$", "", ln)
                # 若包含冒号，取前半
                ln = re.split(r"[：:]", ln)[0].strip()
                return ln
        # 4) 兜底
        return fallback_name
    except Exception:
        return fallback_name


def collect_files(input_path: str) -> List[str]:
    if os.path.isfile(input_path):
        ext = os.path.splitext(input_path)[1].lower()
        if ext in [".txt", ".docx", ".pdf"]:
            return [input_path]
        return []
    files = []
    for root, _, fnames in os.walk(input_path):
        for fn in fnames:
            ext = os.path.splitext(fn)[1].lower()
            if ext in [".txt", ".docx", ".pdf"]:
                files.append(os.path.join(root, fn))
    files.sort()
    return files


def ensure_sheet(wb, sheet_name: str):
    # 统一的基础与扩展表头定义，便于新建或补齐现有工作表
    base_headers = [
        "序号","剧本名","文本/URL","剧本概要","字数(千字)","页数","建议集数区间",
    ]
    extra_headers = {
        # 内容：倾向/时代/主题/导演建议
        "H": "倾向（男频/女频/双频）",
        "I": "时代背景（现代/古代/民国/科幻/其他）",
        "J": "主题标签",
        "K": "导演（风格建议）",
        # 制作：场景/演员
        "L": "制作｜场景数量/内外景占比（8）",
        "M": "制作｜演员数量结构（7）",
        # 宣发：受众/上线/营销
        "N": "宣发｜受众定位",
        "O": "宣发｜上线时间",
        "P": "宣发｜营销方向",
        # 冲突分项
        "Q": "冲突｜类型覆盖",
        "R": "冲突｜升级曲线",
        "S": "冲突｜风险与代价",
        "T": "冲突｜解决策略多样性",
        # 台词分项
        "U": "台词｜信息密度",
        "V": "台词｜金句率与记忆点",
        "W": "台词｜生活化/专业度",
        "X": "台词｜价值观与合规",
        # 剧情结构分项与总分
        "Y": "剧情结构｜起承转合完整度",
        "Z": "剧情结构｜转折密度",
        "AA": "剧情结构｜悬念闭合率",
        "AB": "剧情结构｜节奏均衡度",
        "AC": "剧情结构（总分）",
        # 角色分项与总分
        "AD": "角色塑造｜目标-动机-障碍",
        "AE": "角色塑造｜弧光与反差",
        "AF": "角色塑造｜关系网复杂度",
        "AG": "角色塑造｜记忆点",
        "AH": "角色塑造（总分）",
        # 冲突总分
        "AI": "冲突设置（总分）",
        # 台词总分与副本
        "AJ": "台词｜人设匹配",
        "AK": "台词｜信息密度（副本）",
        "AL": "台词质量（总分）",
        # 人物饱满度四组件与评分/标签
        "AM": "人物饱满度｜关键词权重",
        "AN": "人物饱满度｜连接词权重",
        "AO": "人物饱满度｜反差词对",
        "AP": "人物饱满度｜事件层级与弧光",
        "AQ": "人物饱满度（评分）",
        "AR": "人物饱满度（标签）",
        # 制作/宣发总分
        "AS": "制作（总分）",
        "AT": "宣发｜受众定位（分）",
        "AU": "宣发｜上线时间（分）",
        "AV": "宣发｜营销方向（分）",
        "AW": "宣发（总分）",
        # 商业价值四组件与指数
        "AX": "商业｜题材热度（2）",
        "AY": "商业｜传播因子（1.5）",
        "AZ": "商业｜可生产性（1）",
        "BA": "商业｜合规风险（0.5）",
        "BB": "商业价值指数（0–100）",
        # 总分与等级
        "BC": "总分（0–100）",
        "BD": "评分等级",
        # 平台推荐
        "BH": "平台推荐（渠道）",
    }

    if sheet_name in wb.sheetnames:
        # 已存在工作表：补齐缺失的表头（包括 A~H 与扩展列）
        ws = wb[sheet_name]
        # 基础表头 A~G
        for idx, val in enumerate(base_headers, start=1):
            cell = ws.cell(row=1, column=idx)
            if cell.value in (None, ""):
                cell.value = val
        # 若 D 列仍为“文本粘贴”，统一改为“剧本概要”以匹配规则
        d_cell = ws.cell(row=1, column=4)
        if (d_cell.value or "").strip() != "剧本概要":
            d_cell.value = "剧本概要"
        # 扩展表头（含 H 与 BH 等）
        for col_letter, title in extra_headers.items():
            idx = column_index_from_string(col_letter)
            cell = ws.cell(row=1, column=idx)
            if cell.value in (None, ""):
                cell.value = title
        return ws

    # 不存在工作表：新建并写入完整表头（如可能，从“评估输入”复制）
    headers = None
    if "评估输入" in wb.sheetnames:
        ws_src = wb["评估输入"]
        headers = [ws_src.cell(row=1, column=c).value for c in range(1, ws_src.max_column + 1)]
    ws = wb.create_sheet(title=sheet_name)
    if headers:
        for idx, val in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=val)
    else:
        for idx, val in enumerate(base_headers, start=1):
            ws.cell(row=1, column=idx, value=val)
        for col_letter, title in extra_headers.items():
            idx = column_index_from_string(col_letter)
            ws.cell(row=1, column=idx, value=title)
    return ws


def first_writable_row(ws) -> int:
    """
    返回安全的数据起始行：
    - 至少为第 2 行（第 1 行为表头）。
    - 跳过任何包含在合并单元格范围内的行，避免写入到合并行导致 MergedCell 只读错误。
    - 同时考虑当前已有数据行数（ws.max_row），取三者中的最大值+必要偏移。
    """
    try:
        # 统计所有被合并覆盖的行（包含起止）
        reserved_rows = set()
        try:
            for rng in ws.merged_cells.ranges:
                # openpyxl 的 CellRange 提供 min_row/max_row 属性
                min_r = getattr(rng, 'min_row', None)
                max_r = getattr(rng, 'max_row', None)
                if min_r is None or max_r is None:
                    # 兜底：使用 bounds (min_col, min_row, max_col, max_row)
                    bounds = getattr(rng, 'bounds', None)
                    if bounds and len(bounds) == 4:
                        min_r = bounds[1]
                        max_r = bounds[3]
                if isinstance(min_r, int) and isinstance(max_r, int):
                    for r in range(min_r, max_r + 1):
                        reserved_rows.add(r)
        except Exception:
            reserved_rows = set()

        # 基线：第 2 行开始写入
        base_start = 2
        # 当前数据末行 + 1
        data_start = (ws.max_row or 1) + 1
        # 合并行之后一行
        merge_start = (max(reserved_rows) + 1) if reserved_rows else 2
        start = max(base_start, data_start, merge_start)
        return start
    except Exception:
        # 失败兜底：至少第 2 行
        return 2


def ensure_excel_file(excel_path: str):
    """若目标 Excel 不存在，则创建最小工作簿并写入基础表头。"""
    try:
        if os.path.exists(excel_path):
            return
        from openpyxl import Workbook
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = '工作表1'
        base_headers = [
            "序号","剧本名","文本/URL","剧本概要","字数(千字)","页数","建议集数区间",
        ]
        for idx, val in enumerate(base_headers, start=1):
            ws.cell(row=1, column=idx, value=val)
        wb.save(excel_path)
    except Exception as e:
        # 兜底：至少确保文件存在，避免首次 load_workbook 报错
        try:
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            open(excel_path, 'a').close()
        except Exception:
            pass
        print(f"[警告] Excel 初始化失败：{e}")


def apply_formulas(ws, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        # 建议集数区间（G列）：基于 E 列集数
        ws[f"G{r}"] = f'=IF(E{r}<=8,"15–20集",IF(E{r}<=20,"20–30集","30–40集"))'
        # 剧情结构总分（AC）：Y+Z+AA+AB
        ws[f"AC{r}"] = f'=Y{r}+Z{r}+AA{r}+AB{r}'
        # 角色塑造总分（AH）：AD+AE+AF+AG
        ws[f"AH{r}"] = f'=AD{r}+AE{r}+AF{r}+AG{r}'
        # 冲突设置总分（AI）：Q+R+S+T
        ws[f"AI{r}"] = f'=Q{r}+R{r}+S{r}+T{r}'
        # 台词质量总分（AL）：AJ+U+V+W+X
        ws[f"AL{r}"] = f'=AJ{r}+U{r}+V{r}+W{r}+X{r}'
        # 信息密度副本（AK）：与 U 列保持一致，供台词计算与视图使用
        ws[f"AK{r}"] = f'=U{r}'
        # 制作总分（AS）：场景数量与演员结构，封顶 15 分
        ws[f"AS{r}"] = f'=MIN(15,L{r}+M{r})'
        # 宣发总分（AW）：受众定位/上线时间/营销方向，封顶 10 分
        ws[f"AW{r}"] = f'=MIN(10,AT{r}+AU{r}+AV{r})'
        # 人物饱满度（AQ，0–100）：按 AM/AN/AO/AP 权重 40%/30%/20%/10%
        ws[f"AQ{r}"] = (
            f'=ROUND(AM{r}*0.4 + AN{r}*0.3 + AO{r}*0.2 + AP{r}*0.1,2)'
        )
        # 人物饱满度标签（AR）：按 AQ（0–100）映射 高/中/低
        ws[f"AR{r}"] = f'=IF(AQ{r}>=67,"高",IF(AQ{r}>=34,"中","低"))'
        # 商业价值指数（BB）：AX/AY/AZ/BA 加权归一到 0–100
        ws[f"BB{r}"] = (
            f'=ROUND((AX{r}/2*0.4 + AY{r}/1.5*0.3 + AZ{r}/1*0.2 + BA{r}/0.5*0.1)*100,2)'
        )
        # 总分（BC）：内容四项 + 人物饱满度（AQ/100*20）+ 制作 + 宣发 + 商业价值指数的 5%
        ws[f"BC{r}"] = (
            f'=ROUND(AC{r}+AH{r}+AI{r}+AL{r}+(AQ{r}/100*20)+AS{r}+AW{r}+(BB{r}/100*5),2)'
        )
        # 评分等级（BD）：区间映射
        ws[f"BD{r}"] = (
            f'=IF(BC{r}<=20,"极弱",IF(BC{r}<=40,"偏弱",IF(BC{r}<=60,"中等",IF(BC{r}<=80,"较强","卓越"))))'
        )
        # 平台推荐（BH）：在原有规则基础上，纳入宣发受众（N/AT）、上线时间（O/AU）、营销方向（P/AV）权重。
        # 受众上行偏好（DY/XHS）：
        #   AT>=60 或 AV>=2 或 AU>=2，且 N 含“一二线/18–35/女性/女”，或 P 含“反转/成长/救赎/甜虐/知识”，或 O 含“周末/节假”。
        # 受众下行偏好（RG/KS）：
        #   AT>=60 或 AV>=2 或 AU>=2，且 N 含“下沉/45/男性/男”，或 P 含“家国/叙事/家庭”，或 O 含“工作日”。
        # 映射策略：各 BB 档位下结合 倾向(H) 与上述偏好；双频提供覆盖或双平台组合。
        ws[f"BH{r}"] = (
            f'=IF(BB{r}>=70,'
            f'IF(H{r}="双频","抖音/红果/快手/小红书",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("一二线",N{r})),ISNUMBER(SEARCH("18–35",N{r})),ISNUMBER(SEARCH("女性",N{r})),ISNUMBER(SEARCH("女",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("反转",P{r})),ISNUMBER(SEARCH("成长",P{r})),ISNUMBER(SEARCH("救赎",P{r})),ISNUMBER(SEARCH("甜虐",P{r})),ISNUMBER(SEARCH("知识",P{r})))),'
            f'       AND(AU{r}>=2,OR(ISNUMBER(SEARCH("周末",O{r})),ISNUMBER(SEARCH("节假",O{r}))))),'
            f'   "抖音/小红书/红果",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("下沉",N{r})),ISNUMBER(SEARCH("45",N{r})),ISNUMBER(SEARCH("男性",N{r})),ISNUMBER(SEARCH("男",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("家国",P{r})),ISNUMBER(SEARCH("叙事",P{r})),ISNUMBER(SEARCH("家庭",P{r})))),'
            f'       AND(AU{r}>=2,ISNUMBER(SEARCH("工作日",O{r})))),'
            f'   "红果/快手/抖音",'
            f'IF(H{r}="男频","红果/快手/抖音","抖音/小红书/红果")))),'
            f'IF(BB{r}>=50,'
            f'IF(H{r}="双频","抖音/红果/小红书",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("一二线",N{r})),ISNUMBER(SEARCH("18–35",N{r})),ISNUMBER(SEARCH("女性",N{r})),ISNUMBER(SEARCH("女",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("反转",P{r})),ISNUMBER(SEARCH("成长",P{r})),ISNUMBER(SEARCH("救赎",P{r})),ISNUMBER(SEARCH("甜虐",P{r})),ISNUMBER(SEARCH("知识",P{r})))),'
            f'       AND(AU{r}>=2,OR(ISNUMBER(SEARCH("周末",O{r})),ISNUMBER(SEARCH("节假",O{r}))))),'
            f'   "抖音/小红书",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("下沉",N{r})),ISNUMBER(SEARCH("45",N{r})),ISNUMBER(SEARCH("男性",N{r})),ISNUMBER(SEARCH("男",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("家国",P{r})),ISNUMBER(SEARCH("叙事",P{r})),ISNUMBER(SEARCH("家庭",P{r})))),'
            f'       AND(AU{r}>=2,ISNUMBER(SEARCH("工作日",O{r})))),'
            f'   "红果/快手/抖音",'
            f'IF(H{r}="男频","红果/抖音","抖音/小红书")))),'
            f'IF(BB{r}>=30,'
            f'IF(H{r}="双频","抖音（测试）/红果（小投）/小红书（小投）",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("一二线",N{r})),ISNUMBER(SEARCH("18–35",N{r})),ISNUMBER(SEARCH("女性",N{r})),ISNUMBER(SEARCH("女",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("反转",P{r})),ISNUMBER(SEARCH("成长",P{r})),ISNUMBER(SEARCH("救赎",P{r})),ISNUMBER(SEARCH("甜虐",P{r})),ISNUMBER(SEARCH("知识",P{r})))),'
            f'       AND(AU{r}>=2,OR(ISNUMBER(SEARCH("周末",O{r})),ISNUMBER(SEARCH("节假",O{r}))))),'
            f'   "抖音（测试）/小红书（小投）",'
            f'IF(OR(AND(AT{r}>=60,OR(ISNUMBER(SEARCH("下沉",N{r})),ISNUMBER(SEARCH("45",N{r})),ISNUMBER(SEARCH("男性",N{r})),ISNUMBER(SEARCH("男",N{r})))),'
            f'       AND(AV{r}>=2,OR(ISNUMBER(SEARCH("家国",P{r})),ISNUMBER(SEARCH("叙事",P{r})),ISNUMBER(SEARCH("家庭",P{r})))),'
            f'       AND(AU{r}>=2,ISNUMBER(SEARCH("工作日",O{r})))),'
            f'   "抖音（测试）/红果（小投）/快手（小投）",'
            f'IF(H{r}="男频","抖音（测试）/红果（小投）","抖音（测试）/小红书（小投）")))),'
            f'"抖音（测试）")))'
        )


def _split_sentences(text: str) -> List[str]:
    """将中文文本切分为句子，保留句末标点。"""
    if not text:
        return []
    text = re.sub(r"\s+", " ", text)
    # 保留句末标点的切分
    sents = re.findall(r"[^。！？!?]*[。！？!?]", text)
    if not sents:
        # 兜底：按换行或逗号粗切
        sents = re.split(r"[\n,，]", text)
    # 去掉过短/纯噪声句子
    cleaned = []
    for s in sents:
        s = s.strip()
        if len(s) >= 6:
            cleaned.append(s)
    # 保留更多句子以覆盖全文（上限 4000，避免过长导致内存偏大）
    return cleaned[:4000]


def _detect_style(title: str, text: str) -> Tuple[str, str]:
    """识别摘要风格：返回 (倾向标签, 主题/题材标签)。"""
    # 首选模型方法
    gender = '双频'
    genre = ''
    try:
        if detect_gender_channel:
            g = detect_gender_channel(text, title)
            gender = {'男': '男频', '女': '女频'}.get(g, '双频')
        if detect_genre:
            # 时代用于辅助 genre 检测
            era = '现代'
            if detect_era:
                e = detect_era(text, title)
                era = {'现代':'现代','古代':'古代','玄幻/奇幻':'科幻/奇幻'}.get(e, '现代')
            genre = detect_genre(text, title, era) or ''
    except Exception:
        pass
    # 兜底：基于关键词的启发式
    if not genre:
        if any(k in (title+text) for k in ['系统','逆袭','打脸','商战','科技','权谋']):
            genre = '都市/商战/权谋'
        elif any(k in (title+text) for k in ['甜宠','虐恋','霸总','婚恋','家庭']):
            genre = '甜宠/婚恋/家庭'
        elif any(k in (title+text) for k in ['江湖','武侠','朝堂','官职']):
            genre = '武侠/权谋'
    return gender, genre


def summarize_text(text: str, title: str = "", target_chars: int = 500) -> str:
    """面向全文的 500 字左右剧情概要（风格细分 + 精细权重）。

    设计目标：
    - 覆盖全文关键剧情：从开端/中段/结尾分区均衡选句，而非仅取前500字；
    - 风格细分：依据倾向（男频/女频/双频）与题材，动态调整关键词与权重；
    - 精细分词权重：结合 TF-IDF/TextRank 与词性加权（名词/动词/形容词优先）。
    长度控制在 450–650 字之间。
    """
    try:
        if not text:
            return title or ""
        # 清理版权/模板噪声（不删除正文关键词）
        noise_patterns = [r"版权", r"声明", r"禁止转载", r"作者有话说"]
        for pat in noise_patterns:
            text = re.sub(pat, "", text)
        sents = _split_sentences(text)
        if not sents:
            return (text[:target_chars]).strip()
        # 风格细分：确定倾向/题材标签，构建权重词典
        gender, genre = _detect_style(title, text)
        # 基础关键词权重（按倾向调整）
        base_weights: Dict[str, float] = {
            # 人物与关系
            '主角':1.2,'男主':1.2,'女主':1.2,'反派':1.2,'家庭':0.9,'职场':0.9,'都市':0.8,'乡村':0.8,
            # 冲突与结构
            '系统':1.4,'逆袭':1.4,'打脸':1.4,'复仇':1.3,'成长':1.2,'救赎':1.2,'人性':1.1,
            '权谋':1.3,'悬疑':1.0,'冲突':1.1,'危机':1.0,'高潮':1.1,'结局':1.1,
            # 目标-动机-障碍
            '目标':1.0,'动机':1.0,'障碍':1.0,'代价':1.1,'计划':0.9,'关系':0.9,'阵营':0.9,'挑战':1.0,'改变':1.0,'转折':1.2,
            # 女频倾向
            '甜宠':1.4,'虐恋':1.4,'霸总':1.2,'婚恋':1.1,
        }
        if gender == '男频':
            base_weights.update({'系统':1.6,'逆袭':1.6,'打脸':1.5,'权谋':1.4,'商战':1.3,'科技':1.2})
        elif gender == '女频':
            base_weights.update({'甜宠':1.6,'虐恋':1.6,'霸总':1.4,'家庭':1.1,'治愈':1.1})
        # 题材微调
        if '武侠' in genre or '权谋' in genre:
            base_weights.update({'江湖':1.3,'宫廷':1.2,'官职':1.1,'朝堂':1.2})
        if '商战' in genre:
            base_weights.update({'公司':1.1,'项目':1.0,'融资':1.1,'合约':1.0})

        # 分词与关键词总体权重（TF-IDF/TextRank + 词性加权）
        tfidf_weights: Dict[str, float] = {}
        textrank_weights: Dict[str, float] = {}
        if JIEBA_ANALYSE_AVAILABLE:
            try:
                tfidf = jieba_analyse.extract_tags(text, topK=200, withWeight=True)
                tfidf_weights = {w: float(wt) for w, wt in tfidf}
            except Exception:
                tfidf_weights = {}
            try:
                tr = jieba_analyse.textrank(text, topK=200, withWeight=True)
                textrank_weights = {w: float(wt) for w, wt in tr}
            except Exception:
                textrank_weights = {}

        # 词性加权：名词/动词/形容词优先
        pos_boost: Dict[str, float] = {}
        if JIEBA_AVAILABLE:
            try:
                for wflag in pseg.lcut(text[:200000]):  # 词性统计限制前20万字符避免性能问题
                    w = wflag.word
                    f = wflag.flag
                    if len(w) < 2:
                        continue
                    if f.startswith('n'):  # 名词
                        pos_boost[w] = max(pos_boost.get(w, 0.0), 1.2)
                    elif f.startswith('v'):  # 动词
                        pos_boost[w] = max(pos_boost.get(w, 0.0), 1.1)
                    elif f.startswith('a'):  # 形容词
                        pos_boost[w] = max(pos_boost.get(w, 0.0), 1.05)
            except Exception:
                pos_boost = {}

        # 为每句打分（覆盖全文，不偏前段）
        # 计算句子在全文的相对位置（近似：累计长度占比）
        cum = 0
        total_len = sum(len(s) for s in sents)
        sent_positions = []  # (idx, pos_ratio)
        for i, s in enumerate(sents):
            cum += len(s)
            pos_ratio = cum / max(1, total_len)
            sent_positions.append((i, pos_ratio))

        def sentence_score(i: int, s: str) -> float:
            # 关键词权重
            kw = sum(base_weights.get(k, 0.0) * s.count(k) for k in base_weights.keys())
            # TF-IDF & TextRank 权重
            token_weight = 0.0
            if JIEBA_AVAILABLE:
                try:
                    for tok in jieba.lcut(s):
                        if len(tok) < 2:
                            continue
                        token_weight += tfidf_weights.get(tok, 0.0) + 0.8 * textrank_weights.get(tok, 0.0)
                        token_weight *= pos_boost.get(tok, 1.0)
                except Exception:
                    pass
            # 连接词与弧光加分
            connectors = ['但是','然而','一方面','另一方面','看似','实则','表面','内心']
            arcs = ['反转','黑化','洗白','救赎','成长','挣扎','人性','代价','转折']
            conn_hits = sum(1 for c in connectors if c in s)
            arc_hits = sum(1 for a in arcs if a in s)
            conn_arc = 0.5 * conn_hits + 0.7 * arc_hits
            # 位置均衡：鼓励覆盖中后段
            pos = sent_positions[i][1]
            pos_bonus = 0.6 if 0.25 <= pos <= 0.75 else (0.8 if pos > 0.75 else 0.4)
            # 长度惩罚（过长句不利摘要）
            length_penalty = 0.9 if len(s) > 180 else 1.0
            return kw + token_weight + conn_arc + pos_bonus

        # 对每一分区选取高分句：开端、中段、结尾
        # 将索引按位置分区
        indices = list(range(len(sents)))
        zone_A = [i for i, pos in sent_positions if pos <= 0.25]
        zone_B = [i for i, pos in sent_positions if 0.25 < pos <= 0.75]
        zone_C = [i for i, pos in sent_positions if pos > 0.75]

        def top_sentences(zone: List[int], limit: int) -> List[str]:
            scored = []
            for i in zone:
                s = sents[i].strip()
                if not s or len(s) < 6:
                    continue
                sc = sentence_score(i, s)
                scored.append((sc, i, s))
            scored.sort(key=lambda x: x[0], reverse=True)
            # 选取并截断到适长
            picked = []
            for _, i, s in scored[:limit]:
                if len(s) > 200:
                    s = s[:200]
                picked.append((i, s))
            # 保持原文顺序
            picked.sort(key=lambda x: x[0])
            return [s for _, s in picked]

        # 按目标长度分配各区份比例（约 500 字：A 30%、B 40%、C 30%）
        target = max(450, min(650, target_chars))
        A_parts = top_sentences(zone_A, 5)
        B_parts = top_sentences(zone_B, 6)
        C_parts = top_sentences(zone_C, 5)
        selected: List[str] = []
        total_chars = 0
        def append_parts(parts: List[str]):
            nonlocal total_chars
            for s in parts:
                if total_chars >= target:
                    break
                selected.append(s)
                total_chars += len(s)

        # 先按区份追加，再视剩余长度补充次优句
        append_parts(A_parts)
        append_parts(B_parts)
        append_parts(C_parts)
        if total_chars < int(target * 0.9):
            # 兜底：从全局高分但未选中的句子补齐
            rest_indices = set(indices) - set([i for i,_ in enumerate(sents) if False])
            scored_all = []
            for i in range(len(sents)):
                s = sents[i].strip()
                if not s:
                    continue
                sc = sentence_score(i, s)
                scored_all.append((sc, i, s))
            scored_all.sort(key=lambda x: x[0], reverse=True)
            for _, i, s in scored_all:
                if total_chars >= target:
                    break
                # 避免重复加入区份已选句子
                if s in selected:
                    continue
                if len(s) > 200:
                    s = s[:200]
                selected.append(s)
                total_chars += len(s)

        summary = ''.join(selected)
        prefix = f"《{title}》主要讲述：" if title else "剧情概要："
        combined = (prefix + summary).strip()
        if len(combined) > 650:
            combined = combined[:650]
        # 保证不少于 450 字，如仍不足，追加中段连续句以增强连贯性
        while len(combined) < 450 and B_parts:
            extra = B_parts.pop(0)
            combined += extra
            if len(combined) > 650:
                combined = combined[:650]
                break
        return combined
    except Exception:
        return (text[:target_chars]).strip()


def main():
    parser = argparse.ArgumentParser(description="剧本评估通用导入 CLI")
    parser.add_argument(
        "--input",
        default=".",
        help="输入路径（文件或目录），支持 .txt/.docx/.pdf；默认当前工作目录",
    )
    parser.add_argument(
        "--excel",
        default="/Users/mr.hu/Desktop/爆款排名/剧本评估表.xlsx",
        help="目标 Excel 文件路径",
    )
    parser.add_argument(
        "--sheet",
        default="工作表1",
        help="目标工作表名，不存在则自动创建（默认：工作表1）",
    )
    parser.add_argument(
        "--max",
        type=int,
        default=100,
        help="单次最多处理的文件数（默认100）",
    )
    args = parser.parse_args()

    files = collect_files(args.input)
    if not files:
        print("[提示] 未发现可处理的文件（支持 .txt/.docx/.pdf）。")
        sys.exit(0)

    if len(files) > args.max:
        print(f"[提示] 本次仅处理前 {args.max} 个文件，其余 {len(files) - args.max} 个请分批处理或调整 --max。")
        files = files[:args.max]

    print(f"[信息] 即将处理 {len(files)} 个文件，写入工作表：{args.sheet}")

    # 若目标 Excel 不存在，则自动初始化
    ensure_excel_file(args.excel)
    wb = load_workbook(args.excel)
    ws = ensure_sheet(wb, args.sheet)

    # 使用更稳健的起始行，避免写入到合并行（如模板第二行为合并标题行）。
    start_row = first_writable_row(ws)
    cur_row = start_row
    pdf_failures = []

    for idx, path in enumerate(files, start=1):
        name = os.path.splitext(os.path.basename(path))[0]
        ext = os.path.splitext(path)[1].lower()
        text = ""
        tip = ""
        if ext == ".txt":
            text = read_txt(path)
        elif ext == ".docx":
            text = read_docx(path)
        elif ext == ".pdf":
            text, tip = read_pdf(path)
        else:
            continue

        # 提取真实剧本名（优先文本标题，兜底用文件名）
        title = extract_title(text, name)

        # 写入基础信息
        ws.cell(row=cur_row, column=1, value=cur_row - 1)  # 序号
        ws.cell(row=cur_row, column=2, value=title)        # 剧本名
        ws.cell(row=cur_row, column=3, value=path)         # 文本/URL
        # 文本粘贴改为“500字左右概要”并加入质量校验
        try:
            from summary_quality import generate_summary
        except Exception:
            generate_summary = None
        if generate_summary:
            try:
                summary = generate_summary(text, title, target_chars=500)
            except Exception:
                summary = (text[:500]).strip()
        else:
            try:
                summary = summarize_text(text, title, target_chars=500)
            except Exception:
                summary = (text[:500]).strip()
        ws.cell(row=cur_row, column=4, value=summary)

        # 调用现有评分逻辑填充各项指标
        try:
            fill_row(ws, cur_row, title, text)
        except Exception as e:
            print(f"[警告] 行 {cur_row} 填充失败：{e}")

        if ext == ".pdf" and tip:
            pdf_failures.append((path, tip))

        cur_row += 1

    # 补齐公式联动：为稳健起见，统一对整张工作表(自第2行至当前最大行)回填一次，
    # 可避免已有数据在外部编辑后出现公式缺失的情况。
    end_row = ws.max_row
    apply_formulas(ws, 2, end_row)
    wb.save(args.excel)

    print(f"[完成] 已写入 {end_row - start_row + 1} 行到工作表：{args.sheet}")
    if pdf_failures:
        print("[提示] 以下 PDF 未能解析，请改用 txt/docx：")
        for p, t in pdf_failures:
            print(f" - {p}：{t}")


if __name__ == "__main__":
    main()