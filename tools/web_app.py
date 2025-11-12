#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import time
import shutil
import json
import sqlite3
from typing import List, Optional, Dict

from flask import Flask, request, render_template, redirect, url_for, send_file, session, jsonify
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from uuid import uuid4
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash


BASE_DIR = os.getcwd()
DEFAULT_EXCEL = os.environ.get('EXCEL_PATH') or os.path.join(BASE_DIR, '剧本评估表.xlsx')
DEFAULT_SHEET = os.environ.get('SHEET_NAME') or '工作表1'
UPLOAD_DIR = os.environ.get('UPLOAD_DIR') or os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXT = {'.txt', '.docx', '.pdf'}
DATA_DIR = os.environ.get('DATA_DIR') or BASE_DIR
# 将数据库持久化到挂载的数据目录，避免容器内 /app 写入权限或镜像更新导致的不可用
# 结构：/data/system/nebula.db
DB_PATH = os.path.join(DATA_DIR, 'system', 'nebula.db')
MAX_FILES = int(os.environ.get('MAX_FILES', '100'))

# 允许直接导入 tools 目录下的脚本
TOOLS_DIR = os.path.join(BASE_DIR, 'tools')
if TOOLS_DIR not in sys.path:
    sys.path.insert(0, TOOLS_DIR)

import universal_cli as uni  # 复用读取与公式函数
from auto_score_from_text import fill_row  # 复用评分填充逻辑
try:
    from trending import load_trending_config as _load_trend_cfg
except Exception:
    _load_trend_cfg = None
try:
    import ocr_pipeline as ocrp  # OCR 管线（PDF/图片 -> 文本+HOCR）
except Exception:
    ocrp = None


def resource_path(*parts):
    """统一资源路径解析（移除打包兼容分支）。"""
    return os.path.join(BASE_DIR, *parts)

def ensure_excel_file(excel_path: str):
    """确保 Excel 文件可用：不存在则创建最小工作簿（不依赖打包模板）。"""
    try:
        if os.path.exists(excel_path):
            return
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '工作表1'
        base_headers = [
            '序号','剧本名','文本/URL','剧本概要','字数(千字)','页数','建议集数区间'
        ]
        for idx, val in enumerate(base_headers, start=1):
            ws.cell(row=1, column=idx, value=val)
        wb.save(excel_path)
    except Exception as e:
        # 兜底失败时仅创建空文件，减少首次 load_workbook 出错概率
        try:
            open(excel_path, 'a').close()
        except Exception:
            pass
        print(f"[警告] 默认 Excel 初始化失败: {e}")


# 指定模板目录
TEMPLATES_DIR = os.path.join(BASE_DIR, 'tools', 'templates')
app = Flask(__name__, template_folder=TEMPLATES_DIR)
# 会话隔离所需的密钥（生产环境建议通过环境变量 SECRET_KEY 设置）
app.secret_key = os.environ.get('SECRET_KEY', 'nebula-secret-key-change-me')
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # 如启用 HTTPS 可改为 True
API_KEY = os.environ.get('API_KEY')  # 若设置则开启 API Key 校验（Header: X-API-Key）


# -------------------- 账号与操作留存（SQLite） --------------------

def init_db():
    # 确保数据库所在目录存在
    try:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    except Exception as e:
        print(f"[警告] 创建数据库目录失败: {e}")
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT,
                password_hash TEXT NOT NULL,
                created_at INTEGER NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS operations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                detail TEXT,
                ts INTEGER NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass


def get_user_by_username(username: str) -> Optional[Dict]:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, username, email, password_hash FROM users WHERE username=?", (username,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return None
    return {"id": row[0], "username": row[1], "email": row[2], "password_hash": row[3]}


def get_user_by_id(uid: int) -> Optional[Dict]:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # 兼容可能不存在 is_admin 字段的旧库
    try:
        cur.execute("SELECT id, username, email, is_admin FROM users WHERE id=?", (uid,))
        row = cur.fetchone()
    except Exception:
        cur.execute("SELECT id, username, email FROM users WHERE id=?", (uid,))
        r = cur.fetchone()
        row = (r[0], r[1], r[2], 0) if r else None
    conn.close()
    if not row:
        return None
    return {"id": row[0], "username": row[1], "email": row[2], "is_admin": int(row[3] or 0)}


def get_user_credentials(uid: int) -> Optional[Dict]:
    """获取用户的凭据信息（含密码哈希与管理员标记）。"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, username, password_hash, is_admin FROM users WHERE id=?", (uid,))
        row = cur.fetchone()
    except Exception:
        cur.execute("SELECT id, username, password_hash FROM users WHERE id=?", (uid,))
        r = cur.fetchone()
        row = (r[0], r[1], r[2], 0) if r else None
    conn.close()
    if not row:
        return None
    return {"id": row[0], "username": row[1], "password_hash": row[2], "is_admin": int(row[3] or 0)}


def create_user(username: str, password: str, email: str = "") -> int:
    pwd_hash = generate_password_hash(password)
    ts = int(time.time())
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO users (username, email, password_hash, created_at) VALUES (?, ?, ?, ?)",
        (username, email, pwd_hash, ts),
    )
    conn.commit()
    uid = cur.lastrowid
    conn.close()
    return uid


def log_action(action: str, detail: dict = None):
    uid = session.get('user_id')
    if not uid:
        return
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO operations (user_id, action, detail, ts) VALUES (?, ?, ?, ?)",
        (uid, action, json.dumps(detail or {}), int(time.time())),
    )
    conn.commit()
    conn.close()


def update_password(uid: int, new_password: str) -> bool:
    """更新用户密码哈希。"""
    try:
        pwd_hash = generate_password_hash(new_password)
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("UPDATE users SET password_hash=? WHERE id=?", (pwd_hash, uid))
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def current_user() -> Optional[Dict]:
    uid = session.get('user_id')
    if not uid:
        return None
    return get_user_by_id(int(uid))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('user_id'):
            nxt = request.path
            return redirect(url_for('login', next=nxt))
        return fn(*args, **kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))
        u = current_user()
        if not u or int(u.get('is_admin') or 0) != 1:
            return "需要管理员权限", 403
        return fn(*args, **kwargs)
    return wrapper


def allowed_file(filename: str) -> bool:
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXT

def allowed_ocr_file(filename: str) -> bool:
    """OCR 上传允许的文件类型（PDF 与主流图片）。"""
    ext = os.path.splitext(filename)[1].lower()
    return ext in {'.pdf', '.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp'}


def ensure_upload_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def get_session_id() -> str:
    sid = session.get('sid')
    if not sid:
        sid = uuid4().hex
        session['sid'] = sid
    return sid


def get_session_dir() -> str:
    # 优先使用账号ID实现账号级隔离；未登录则退回会话ID
    uid = session.get('user_id')
    base_id = str(uid) if uid else get_session_id()
    # 将会话数据持久化到可挂载的数据目录（容器中默认 /data）
    d = os.path.join(DATA_DIR, 'user_data', base_id)
    os.makedirs(d, exist_ok=True)
    return d


def ensure_user_upload_dir() -> str:
    uid = session.get('user_id')
    base_id = str(uid) if uid else get_session_id()
    d = os.path.join(UPLOAD_DIR, base_id)
    os.makedirs(d, exist_ok=True)
    return d

def get_ocr_output_dir() -> str:
    """返回当前账号/会话的 OCR 输出根目录（包含 text/hocr/reports/logs）。"""
    d = os.path.join(get_session_dir(), 'ocr')
    os.makedirs(d, exist_ok=True)
    for sub in ('text','hocr','reports','logs'):
        os.makedirs(os.path.join(d, sub), exist_ok=True)
    return d


def session_excel_path() -> str:
    """为当前会话/账号提供独立的 Excel 文件路径并确保存在。"""
    user_dir = get_session_dir()
    path = os.path.join(user_dir, '剧本评估表.xlsx')
    ensure_excel_file(path)
    session['excel_path'] = path
    return path

def new_excel_path_for_upload() -> str:
    """每次上传生成一个全新的评估文件，并按命名规范命名。

    命名：用户名+评估时间+"剧本分析.xlsx"，例如：user123_20240101_剧本分析.xlsx。
    为保证唯一性，时间精度采用到秒：YYYYMMDD_HHMMSS。
    文件存放于当前账号/会话目录下的 evaluations 子目录，确保跨用户严格隔离。
    """
    try:
        user_dir = get_session_dir()
        os.makedirs(os.path.join(user_dir, 'evaluations'), exist_ok=True)
        u = current_user() or {}
        username = str(u.get('username') or u.get('email') or u.get('id') or 'guest')
        ts = time.strftime('%Y%m%d_%H%M%S')
        fname = f"{username}_{ts}_剧本分析.xlsx"
        path = os.path.join(user_dir, 'evaluations', fname)
        ensure_excel_file(path)
        session['excel_path'] = path
        return path
    except Exception as e:
        # 兜底：如新路径生成失败，仍回退到会话级Excel，避免请求失败
        print(f"[警告] 新建评估文件失败，回退到会话Excel：{e}")
        return session_excel_path()

def api_protected(fn):
    """若设置了环境变量 API_KEY，则要求请求头 X-API-Key 校验通过。"""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if API_KEY:
            provided = request.headers.get('X-API-Key')
            if not provided or provided != API_KEY:
                return (json.dumps({
                    'ok': False,
                    'error': 'unauthorized',
                    'message': '缺少或错误的 API Key'
                }, ensure_ascii=False), 401, {'Content-Type': 'application/json'})
        return fn(*args, **kwargs)
    return wrapper


def process_files(files: List[str], excel_path: str, sheet_name: str):
    wb = load_workbook(excel_path)
    ws = uni.ensure_sheet(wb, sheet_name)
    # 使用通用的安全起始行计算，避免写入到模板中的合并行（如第2行分组标题）
    try:
        start_row = uni.first_writable_row(ws)
    except Exception:
        start_row = ws.max_row + 1 if ws.max_row >= 2 else 2
    cur_row = start_row
    count = 0
    pdf_failures = []

    # 单次最多处理 MAX_FILES 个，超出部分提示并忽略
    if len(files) > MAX_FILES:
        print(f"[提示] 本次仅处理前 {MAX_FILES} 个文件，其余 {len(files) - MAX_FILES} 个请分批上传。")
    files_to_process = files[:MAX_FILES]

    for path in files_to_process:
        name = os.path.splitext(os.path.basename(path))[0]
        ext = os.path.splitext(path)[1].lower()
        text = ''
        tip = ''
        if ext == '.txt':
            text = uni.read_txt(path)
        elif ext == '.docx':
            text = uni.read_docx(path)
        elif ext == '.pdf':
            text, tip = uni.read_pdf(path)
        if not text and ext == '.pdf':
            # PDF 解析失败则跳过并记录失败名单，用于界面提示
            print(f"[提示] PDF 解析失败：{path}，请提供 txt/docx 版本")
            pdf_failures.append(os.path.basename(path))
            continue

        # 提取真实剧本名（优先文本标题，兜底用文件名）
        title = uni.extract_title(text, name)

        # 基础字段
        ws.cell(row=cur_row, column=1, value=cur_row - 1)  # 序号
        ws.cell(row=cur_row, column=2, value=title)        # 剧本名
        ws.cell(row=cur_row, column=3, value=path)         # 文本/URL
        # 生成约500字的“剧本概要”，并进行质量校验
        try:
            from summary_quality import generate_summary
        except Exception:
            generate_summary = None

        if generate_summary:
            try:
                summary = generate_summary(text, title, target_chars=500)
            except Exception:
                summary = (text[:500]).strip()
        else:
            # 回退：直接截取前500字
            summary = (text[:500]).strip()

        ws.cell(row=cur_row, column=4, value=summary)      # 剧本概要

        try:
            fill_row(ws, cur_row, title, text)
        except Exception as e:
            print(f"[警告] 行 {cur_row} 填充失败：{e}，文本长度={len(text)}")

        # 只为当前行写入联动公式
        uni.apply_formulas(ws, cur_row, cur_row)

        cur_row += 1
        count += 1

    wb.save(excel_path)
    return count, pdf_failures


# -------------------- 基础健康检查 --------------------

@app.route('/api/health', methods=['GET'])
@api_protected
def api_health():
    return {
        'ok': True,
        'app': 'nebula-ocr',
        'time': int(time.time()),
        'features': {
            'excel': True,
            'ocr': bool(ocrp is not None),
        }
    }


# -------------------- OCR 云端接口 --------------------

@app.route('/api/ocr/upload', methods=['POST'])
@api_protected
def api_ocr_upload():
    if ocrp is None:
        return (json.dumps({'ok': False, 'error': 'ocr_not_available', 'message': 'OCR 模块不可用，请检查依赖安装'}, ensure_ascii=False), 500, {'Content-Type': 'application/json'})

    lang = request.form.get('lang', 'chi_sim+eng')
    try:
        threshold = float(request.form.get('threshold', '0.95'))
    except Exception:
        threshold = 0.95
    try:
        workers = int(request.form.get('workers', '0'))
        workers = workers if workers > 0 else None
    except Exception:
        workers = None

    in_dir = ensure_user_upload_dir()
    saved = []
    files = request.files.getlist('files')
    if not files:
        return (json.dumps({'ok': False, 'error': 'no_files', 'message': '未收到上传文件（支持 PDF/JPG/PNG/TIFF/BMP）'}, ensure_ascii=False), 400, {'Content-Type': 'application/json'})
    for f in files:
        filename = secure_filename(f.filename)
        if not filename or not allowed_ocr_file(filename):
            continue
        ts = int(time.time()*1000)
        out_path = os.path.join(in_dir, f"{ts}_{filename}")
        f.save(out_path)
        saved.append(out_path)

    if not saved:
        return (json.dumps({'ok': False, 'error': 'no_valid_files', 'message': '未选择有效文件（仅支持 PDF/JPG/PNG/TIFF/BMP）'}, ensure_ascii=False), 400, {'Content-Type': 'application/json'})

    out_dir = get_ocr_output_dir()
    try:
        results = ocrp.process_files(saved, out_dir, provider='tesseract', lang=lang, accuracy_threshold=threshold, max_workers=workers)
        success = sum(1 for r in results if r.status == 'success')
        low = sum(1 for r in results if r.status == 'low_accuracy')
        failed = sum(1 for r in results if r.status == 'failed')
        summary = {'success': success, 'low_accuracy': low, 'failed': failed}
        resp = {
            'ok': True,
            'count': len(results),
            'summary': summary,
            'output': {
                'base': out_dir,
                'text': os.path.join(out_dir, 'text'),
                'hocr': os.path.join(out_dir, 'hocr'),
                'reports': os.path.join(out_dir, 'reports'),
                'logs': os.path.join(out_dir, 'logs'),
            },
            'results': [
                {
                    'original_filename': r.original_filename,
                    'file_type': r.file_type,
                    'status': r.status,
                    'text_output_path': r.text_output_path,
                    'accuracy': r.accuracy,
                    'page_count': r.page_count,
                    'duration_seconds': r.duration_seconds,
                    'errors': r.errors,
                } for r in results
            ]
        }
        return (json.dumps(resp, ensure_ascii=False), 200, {'Content-Type': 'application/json'})
    except Exception as e:
        return (json.dumps({'ok': False, 'error': 'ocr_failed', 'message': str(e)}, ensure_ascii=False), 500, {'Content-Type': 'application/json'})


@app.route('/api/ocr/summary', methods=['GET'])
@api_protected
def api_ocr_summary():
    out_dir = get_ocr_output_dir()
    summary_path = os.path.join(out_dir, 'reports', 'ocr_run_summary.json')
    if not os.path.exists(summary_path):
        return (json.dumps({'ok': False, 'error': 'not_found', 'message': '未找到本会话的 OCR 汇总'}, ensure_ascii=False), 404, {'Content-Type': 'application/json'})
    try:
        with open(summary_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return (json.dumps({'ok': True, 'summary': data}, ensure_ascii=False), 200, {'Content-Type': 'application/json'})
    except Exception as e:
        return (json.dumps({'ok': False, 'error': 'read_failed', 'message': str(e)}, ensure_ascii=False), 500, {'Content-Type': 'application/json'})


@app.route('/', methods=['GET'])
def index():
    # 默认首页切换为品牌页“星河无限”
    excel_path = session.get('excel_path') or session_excel_path()
    sheet_name = request.args.get('sheet', DEFAULT_SHEET)
    ensure_excel_file(excel_path)
    return render_template('nebula.html', excel_path=excel_path, sheet_name=sheet_name, user=current_user())


@app.route('/nebula', methods=['GET'])
def nebula_home():
    """品牌首页：星河无限。提供引导与跳转。"""
    excel_path = session.get('excel_path') or session_excel_path()
    sheet_name = request.args.get('sheet', DEFAULT_SHEET)
    ensure_excel_file(excel_path)
    return render_template('nebula.html', excel_path=excel_path, sheet_name=sheet_name, user=current_user())


@app.route('/app', methods=['GET'])
@login_required
def app_upload_form():
    """上传入口（原首页上传页）。"""
    excel_path = session.get('excel_path') or session_excel_path()
    sheet_name = request.args.get('sheet', DEFAULT_SHEET)
    ensure_excel_file(excel_path)
    return render_template('analysis.html', excel_path=excel_path, sheet_name=sheet_name, user=current_user(), max_files=MAX_FILES)


@app.route('/analysis', methods=['GET'])
@login_required
def analysis_page():
    """剧本分析独立页面。"""
    excel_path = session.get('excel_path') or session_excel_path()
    sheet_name = request.args.get('sheet', DEFAULT_SHEET)
    ensure_excel_file(excel_path)
    return render_template('analysis.html', excel_path=excel_path, sheet_name=sheet_name, user=current_user(), max_files=MAX_FILES)


@app.route('/upload', methods=['POST'])
@login_required
def upload():
    try:
        # 每次上传生成全新评估文件（严格隔离，不在旧文件上追加）
        excel_path = new_excel_path_for_upload()
        sheet_name = request.form.get('sheet_name', DEFAULT_SHEET)
        # 使用会话独立上传目录
        user_upload_dir = ensure_user_upload_dir()
        ensure_excel_file(excel_path)

        saved_paths = []
        files = request.files.getlist('files')
        if not files:
            return "未收到上传文件（请选择 .txt/.docx/.pdf 或文件夹）", 400
        for f in files:
            filename = secure_filename(f.filename)
            if not filename or not allowed_file(filename):
                continue
            ts = int(time.time()*1000)
            base = os.path.basename(filename)
            out_path = os.path.join(user_upload_dir, f"{ts}_{base}")
            f.save(out_path)
            saved_paths.append(out_path)

        if not saved_paths:
            return "未选择有效文件（仅支持 .txt/.docx/.pdf）", 400

        count, pdf_failures = process_files(saved_paths, excel_path, sheet_name)
        # 将失败名单写入会话，供预览页提示
        session['pdf_failures'] = pdf_failures
        log_action('upload', {'files': saved_paths, 'count': count, 'sheet': sheet_name, 'pdf_failures': len(pdf_failures)})
        # 预览与下载均读取当前会话的 Excel，不再依赖外部传入路径
        return redirect(url_for('preview', sheet=sheet_name, added=count))
    except Exception as e:
        print(f"[错误] 上传处理失败：{e}")
        return f"处理失败：{e}", 500


def sheet_to_rows(excel_path: str, sheet_name: str, max_rows: int = 50, max_cols: int = 30):
    # 注意：openpyxl 不计算公式，预览将显示已有值或公式文本
    wb = load_workbook(excel_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return ["工作表不存在"], []
    ws = wb[sheet_name]
    n_rows = min(ws.max_row, max_rows)
    n_cols = min(ws.max_column, max_cols)
    header = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
    data = []
    for r in range(2, n_rows + 1):
        row = []
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            # 如果是公式，展示公式文本
            v = cell.value
            row.append(v)
        data.append(row)
    return header, data


@app.route('/preview', methods=['GET'])
@login_required
def preview():
    excel_path = session.get('excel_path') or session_excel_path()
    sheet_name = request.args.get('sheet', DEFAULT_SHEET)
    added = int(request.args.get('added', '0'))
    ensure_excel_file(excel_path)
    # 扩展列数以覆盖评分指标（含到 AY 列）
    header, data = sheet_to_rows(excel_path, sheet_name, max_rows=50, max_cols=60)
    log_action('preview', {'sheet': sheet_name, 'rows': len(data)})
    pdf_fails = session.pop('pdf_failures', [])
    # 舆情快照：在预览页展示当前使用的趋势依据
    trend_cfg = {}
    try:
        trend_cfg = _load_trend_cfg() if _load_trend_cfg else {}
    except Exception:
        trend_cfg = {}

    # 规则对齐摘要：基于已填充的分项指标进行合规与达标提示
    def _num(v, default=0.0):
        try:
            return float(v)
        except Exception:
            return default

    compliance_rows = []
    for i, row in enumerate(data):
        # 标题优先取“剧本名”（B 列），否则用序号或占位
        title = row[1] if len(row) > 1 and row[1] else (f"第{i+1}行")
        # 文本（用于原始指标的提示展示，不写回Excel）
        text_val = row[3] if len(row) > 3 and row[3] else ''
        # 指标取值（0-based 索引）
        # 结构
        s_integrity = _num(row[24]) if len(row) > 24 else 0  # Y
        s_turns = _num(row[25]) if len(row) > 25 else 0       # Z
        s_closure = _num(row[26]) if len(row) > 26 else 0     # AA
        s_rhythm = _num(row[27]) if len(row) > 27 else 0      # AB
        # 角色
        r_gmo = _num(row[29]) if len(row) > 29 else 0         # AD
        r_arc = _num(row[30]) if len(row) > 30 else 0         # AE
        r_net = _num(row[31]) if len(row) > 31 else 0         # AF
        r_mem = _num(row[32]) if len(row) > 32 else 0         # AG
        # 冲突
        c_types = _num(row[16]) if len(row) > 16 else 0       # Q
        c_curve = _num(row[17]) if len(row) > 17 else 0       # R
        c_risk = _num(row[18]) if len(row) > 18 else 0        # S
        c_solutions = _num(row[19]) if len(row) > 19 else 0   # T
        # 台词
        d_match = _num(row[35]) if len(row) > 35 else 0       # AJ
        d_info = _num(row[20]) if len(row) > 20 else 0        # U
        d_quotes = _num(row[21]) if len(row) > 21 else 0      # V
        d_prof = _num(row[22]) if len(row) > 22 else 0        # W
        d_compliance = _num(row[23]) if len(row) > 23 else 0  # X
        # 人物饱满度（组件汇总）
        cf_kw = _num(row[38]) if len(row) > 38 else 0         # AM
        cf_conn = _num(row[39]) if len(row) > 39 else 0       # AN
        cf_pair = _num(row[40]) if len(row) > 40 else 0       # AO
        cf_event = _num(row[41]) if len(row) > 41 else 0      # AP
        fullness_score = round(cf_kw*0.4 + cf_conn*0.3 + cf_pair*0.2 + cf_event*0.1, 2)
        fullness_label = ('高' if fullness_score >= 67 else ('中' if fullness_score >= 34 else '低'))

        # 原始指标（仅用于提示）
        try:
            import re, statistics
            trig_words = ['反转','但','但是','然而','却','其实','看似','结果','谁知','出乎意料','意外']
            trig_hits = sum(text_val.count(t) for t in trig_words)
            twist_per_5k = trig_hits / max(1.0, (len(text_val) / 5000.0))
            foreshadow = ['伏笔','悬念','谜团','未解释线索','埋线']
            payoff = ['解释','揭示','真相','回收','说明','交代']
            f_cnt = sum(text_val.count(w) for w in foreshadow)
            p_cnt = sum(text_val.count(w) for w in payoff)
            closure_ratio = (p_cnt / f_cnt) if f_cnt > 0 else 0.0
            paras = [len(p.strip()) for p in str(text_val).split('\n') if str(p).strip()]
            if len(paras) >= 6:
                mean = statistics.mean(paras)
                stdev = statistics.pstdev(paras)
                cv_ratio = (stdev / mean) if mean > 0 else 0.0
            else:
                cv_ratio = 0.0
            info_tokens = 0
            info_tokens += len(re.findall(r"\d+", str(text_val)))
            info_tokens += sum(str(text_val).count(w) for w in ['公司','项目','计划','目标','方案','事件','案','村','县','市','省','学校','医院','职位'])
            info_density_100 = info_tokens / max(1, (len(str(text_val)) / 100.0))
            thousand = max(1, len(str(text_val)) / 1000.0)
            punch_per_1000 = (str(text_val).count('！') + str(text_val).count('“')) / thousand
        except Exception:
            twist_per_5k = 0.0
            closure_ratio = 0.0
            cv_ratio = 0.0
            info_density_100 = 0.0
            punch_per_1000 = 0.0

        # 等级标签：ok/warn/fail
        def lvl(pass_ok: bool, near: bool) -> str:
            if pass_ok:
                return 'ok'
            return 'warn' if near else 'fail'

        checks = [
            { 'name': '起承转合完整度≥4', 'val': s_integrity, 'pass': s_integrity >= 4, 'hint': '标志词≥4：引子/冲突爆发/反转/高潮/结局', 'level': lvl(s_integrity >= 4, s_integrity == 3) },
            { 'name': '转折密度≥3', 'val': s_turns, 'pass': s_turns >= 3, 'hint': f'每4k–6k字≥1次；当前每5k字≈{round(twist_per_5k,2)}', 'level': lvl(s_turns >= 3, s_turns == 2) },
            { 'name': '悬念闭合率≥2', 'val': s_closure, 'pass': s_closure >= 2, 'hint': f'回收比例≥0.6；当前≈{round(closure_ratio,2)}', 'level': lvl(s_closure >= 2, s_closure == 1) },
            { 'name': '节奏均衡度==2', 'val': s_rhythm, 'pass': s_rhythm == 2, 'hint': f'段落长度CV在0.25–0.8；当前≈{round(cv_ratio,2)}', 'level': lvl(s_rhythm == 2, s_rhythm == 1) },
            { 'name': '目标-动机-障碍≥4', 'val': r_gmo, 'pass': r_gmo >= 4, 'hint': '主反派目标-动机-障碍齐全', 'level': lvl(r_gmo >= 4, r_gmo == 3) },
            { 'name': '弧光与反差≥4', 'val': r_arc, 'pass': r_arc >= 4, 'hint': '反转/黑化/救赎等共现频高、阶段变化明确', 'level': lvl(r_arc >= 4, r_arc == 3) },
            { 'name': '关系网复杂度≥2', 'val': r_net, 'pass': r_net >= 2, 'hint': '关键关系≥4条且关系变化≥2次', 'level': lvl(r_net >= 2, r_net == 1) },
            { 'name': '记忆点≥1', 'val': r_mem, 'pass': r_mem >= 1, 'hint': '昵称/口头禅/行为记忆点≥2', 'level': lvl(r_mem >= 1, r_mem == 0) },
            { 'name': '类型覆盖≥2类', 'val': c_types, 'pass': c_types >= 3, 'hint': '至少覆盖两类（人-人/人-环境/人-自我）', 'level': lvl(c_types >= 3, c_types == 1) },
            { 'name': '升级曲线≥2阶', 'val': c_curve, 'pass': c_curve >= 2, 'hint': '强度递增≥2–3阶（升级/更大/逐步等）', 'level': lvl(c_curve >= 2, c_curve == 1) },
            { 'name': '风险与代价≥2', 'val': c_risk, 'pass': c_risk >= 2, 'hint': '出现高风险词并兑现代价', 'level': lvl(c_risk >= 2, c_risk == 1) },
            { 'name': '解决策略≥1且非神降', 'val': c_solutions, 'pass': c_solutions >= 1, 'hint': '方案≥2种且非神降解题', 'level': lvl(c_solutions >= 1, c_solutions == 0) },
            { 'name': '人设匹配≥2', 'val': d_match, 'pass': d_match >= 2, 'hint': '语体一致、口吻贴合', 'level': lvl(d_match >= 2, d_match == 1) },
            { 'name': '信息密度≥2', 'val': d_info, 'pass': d_info >= 2, 'hint': f'每100字信息点≥1.0–1.5；当前≈{round(info_density_100,2)}', 'level': lvl(d_info >= 2, d_info == 1) },
            { 'name': '金句率≥1', 'val': d_quotes, 'pass': d_quotes >= 1, 'hint': f'每千字≥0.8–2.0；当前≈{round(punch_per_1000,2)}', 'level': lvl(d_quotes >= 1, d_quotes == 0) },
            { 'name': '生活化/专业度=1', 'val': d_prof, 'pass': d_prof >= 1, 'hint': '常识与术语准确', 'level': lvl(d_prof >= 1, d_prof == 0) },
            { 'name': '价值观与合规=1', 'val': d_compliance, 'pass': d_compliance >= 1, 'hint': '无不当话术与尺度', 'level': lvl(d_compliance >= 1, d_compliance == 0) },
        ]

        compliance_rows.append({
            'title': title,
            'fullness_score': fullness_score,
            'fullness_label': fullness_label,
            'checks': checks,
            'raw': {
                'twist_per_5k': round(twist_per_5k, 2),
                'closure_ratio': round(closure_ratio, 2),
                'cv': round(cv_ratio, 2),
                'info_density_100': round(info_density_100, 2),
                'punch_per_1000': round(punch_per_1000, 2),
            }
        })
    return render_template(
        'preview.html',
        excel_path=excel_path,
        sheet_name=sheet_name,
        added=added,
        header=header,
        data=data,
        user=current_user(),
        pdf_fails=pdf_fails,
        failed_count=len(pdf_fails),
        trend_cfg=trend_cfg,
        compliance_rows=compliance_rows,
    )


@app.route('/download', methods=['GET'])
@login_required
def download():
    excel_path = session.get('excel_path') or session_excel_path()
    ensure_excel_file(excel_path)
    fname = os.path.basename(excel_path)
    log_action('download', {'file': fname})
    return send_file(excel_path, as_attachment=True, download_name=fname)


# -------------------- 认证路由 --------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    init_db()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        nxt = request.args.get('next') or url_for('app_upload_form')
        user = get_user_by_username(username)
        if not user or not check_password_hash(user['password_hash'], password):
            return render_template('login.html', error='用户名或密码错误', user=current_user(), next=request.args.get('next', ''))
        session.clear()
        session['user_id'] = int(user['id'])
        session_excel_path()
        log_action('login', {'username': username})
        return redirect(nxt)
    return render_template('login.html', user=current_user(), next=request.args.get('next', ''))


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """管理员登录入口：仅允许 is_admin=1 的账号登录成功。"""
    init_db()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        nxt = request.args.get('next') or url_for('admin_ops')
        user = get_user_by_username(username)
        if not user:
            return render_template('admin_login.html', error='账号不存在或非管理员', user=current_user(), next=request.args.get('next', ''))
        # 读取 is_admin
        creds = get_user_credentials(int(user['id']))
        if int(creds.get('is_admin') or 0) != 1:
            return render_template('admin_login.html', error='账号不存在或非管理员', user=current_user(), next=request.args.get('next', ''))
        if not check_password_hash(creds['password_hash'], password):
            return render_template('admin_login.html', error='用户名或密码错误', user=current_user(), next=request.args.get('next', ''))
        session.clear()
        session['user_id'] = int(user['id'])
        session_excel_path()
        log_action('login', {'username': username, 'role': 'admin'})
        return redirect(nxt)
    return render_template('admin_login.html', user=current_user(), next=request.args.get('next', ''))


@app.route('/register', methods=['GET', 'POST'])
def register():
    init_db()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        email = request.form.get('email', '').strip()
        if not username or not password:
            return render_template('register.html', error='用户名与密码必填', user=current_user())
        if get_user_by_username(username):
            return render_template('register.html', error='用户名已存在', user=current_user())
        uid = create_user(username, password, email)
        session.clear()
        session['user_id'] = int(uid)
        session_excel_path()
        log_action('register', {'username': username, 'email': email})
        return redirect(url_for('app_upload_form'))
    return render_template('register.html', user=current_user())


@app.route('/logout', methods=['GET'])
def logout():
    # 先记录操作再清理会话
    log_action('logout')
    session.clear()
    return redirect(url_for('login'))

@app.route('/admin/users', methods=['GET'])
@login_required
@admin_required
def admin_users():
    init_db()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, username, email, created_at FROM users WHERE is_admin=1 ORDER BY id ASC")
        admins = cur.fetchall()
    except Exception:
        # 兼容旧库：无 is_admin 字段时返回空
        admins = []
    conn.close()
    rows = [
        {
            'id': r[0], 'username': r[1], 'email': r[2] or '',
            'created_at': r[3],
            'time_str': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(r[3]))) if r[3] else ''
        } for r in admins
    ]
    return render_template('admin_users.html', rows=rows, user=current_user())


@app.route('/admin/password', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_password():
    init_db()
    u = current_user()
    if request.method == 'POST':
        current_pwd = request.form.get('current_password', '')
        new_pwd = request.form.get('new_password', '')
        confirm_pwd = request.form.get('confirm_password', '')
        creds = get_user_credentials(int(u['id']))
        if not check_password_hash(creds['password_hash'], current_pwd):
            return render_template('admin_password.html', user=u, error='当前密码不正确')
        if len(new_pwd) < 6:
            return render_template('admin_password.html', user=u, error='新密码至少 6 位')
        if new_pwd != confirm_pwd:
            return render_template('admin_password.html', user=u, error='两次输入不一致')
        if update_password(int(u['id']), new_pwd):
            log_action('change_password', {'for': u['username'], 'role': 'admin'})
            return render_template('admin_password.html', user=current_user(), success='密码已更新')
        else:
            return render_template('admin_password.html', user=u, error='密码更新失败')
    return render_template('admin_password.html', user=u)

@app.route('/admin/ops', methods=['GET'])
@login_required
@admin_required
def admin_ops():
    init_db()
    # 过滤参数
    user_id = request.args.get('user_id', '').strip()
    uid = int(user_id) if user_id.isdigit() else None
    now = int(time.time())
    # 支持 YYYY-MM-DD 格式
    def _parse_date(date_str: str, default_ts: int) -> int:
        try:
            if not date_str:
                return default_ts
            tm = time.strptime(date_str, '%Y-%m-%d')
            return int(time.mktime(tm))
        except Exception:
            return default_ts
    start_ts = _parse_date(request.args.get('start', ''), now - 7*24*3600)
    end_ts = _parse_date(request.args.get('end', ''), now)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    sql = (
        "SELECT o.id, o.user_id, u.username, o.action, o.detail, o.ts "
        "FROM operations o LEFT JOIN users u ON o.user_id = u.id "
        "WHERE o.ts BETWEEN ? AND ?"
    )
    params = [start_ts, end_ts]
    if uid:
        sql += " AND o.user_id = ?"
        params.append(uid)
    sql += " ORDER BY o.ts DESC LIMIT 200"
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    conn.close()

    # 汇总统计
    total_ops = len(rows)
    upload_count = sum(1 for r in rows if r[3] == 'upload')
    login_count = sum(1 for r in rows if r[3] == 'login')
    users_active = len({r[1] for r in rows})

    # 解析 detail JSON
    parsed_rows = []
    for r in rows:
        try:
            detail = json.loads(r[4]) if r[4] else {}
        except Exception:
            detail = {"raw": r[4]}
        parsed_rows.append({
            'id': r[0], 'user_id': r[1], 'username': r[2] or '',
            'action': r[3], 'detail': detail, 'ts': r[5],
            'time_str': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(r[5])))
        })

    return render_template(
        'admin_ops.html',
        rows=parsed_rows,
        total_ops=total_ops,
        upload_count=upload_count,
        login_count=login_count,
        users_active=users_active,
        start=request.args.get('start', ''),
        end=request.args.get('end', ''),
        filter_user_id=user_id or ''
    )

@app.route('/admin/ops.json', methods=['GET'])
@login_required
@admin_required
def admin_ops_json():
    init_db()
    user_id = request.args.get('user_id', '').strip()
    uid = int(user_id) if user_id.isdigit() else None
    now = int(time.time())
    def _parse_date(date_str: str, default_ts: int) -> int:
        try:
            if not date_str:
                return default_ts
            tm = time.strptime(date_str, '%Y-%m-%d')
            return int(time.mktime(tm))
        except Exception:
            return default_ts
    start_ts = _parse_date(request.args.get('start', ''), now - 7*24*3600)
    end_ts = _parse_date(request.args.get('end', ''), now)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    sql = (
        "SELECT o.id, o.user_id, u.username, o.action, o.detail, o.ts "
        "FROM operations o LEFT JOIN users u ON o.user_id = u.id "
        "WHERE o.ts BETWEEN ? AND ?"
    )
    params = [start_ts, end_ts]
    if uid:
        sql += " AND o.user_id = ?"
        params.append(uid)
    sql += " ORDER BY o.ts DESC LIMIT 200"
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    conn.close()

    total_ops = len(rows)
    upload_count = sum(1 for r in rows if r[3] == 'upload')
    login_count = sum(1 for r in rows if r[3] == 'login')
    users_active = len({r[1] for r in rows})
    parsed_rows = []
    for r in rows:
        try:
            detail = json.loads(r[4]) if r[4] else {}
        except Exception:
            detail = {"raw": r[4]}
        parsed_rows.append({
            'id': r[0], 'user_id': r[1], 'username': r[2] or '',
            'action': r[3], 'detail': detail, 'ts': r[5],
            'time_str': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(r[5])))
        })

    return jsonify({
        'rows': parsed_rows,
        'summary': {
            'total_ops': total_ops,
            'upload_count': upload_count,
            'login_count': login_count,
            'users_active': users_active,
            'start': request.args.get('start', ''),
            'end': request.args.get('end', ''),
            'filter_user_id': user_id or ''
        }
    })


if __name__ == '__main__':
    init_db()
    host = os.environ.get('HOST', '127.0.0.1')
    port = int(os.environ.get('PORT', '5000'))
    app.run(host=host, port=port, debug=False)