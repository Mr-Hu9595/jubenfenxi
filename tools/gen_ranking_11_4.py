#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SRC_PATH = "/Users/mr.hu/Desktop/爆款排名/11.4 评估表.xlsx"
SRC_SHEET = "11.4 评估输入"
RANK_SHEET = "11.4 排名"


def read_sheet(ws) -> List[Dict[str, Any]]:
    header = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    name_to_idx = {name: i for i, name in enumerate(header)}

    required = [
        '序号', '剧本名', '总分', '评分等级', '商业价值指数', '倾向', '时代背景', '主题',
        '字数（千字）', '建议集数', '平台推荐', '营销方向', '上线时间', '受众定位', '备注'
    ]
    for k in required:
        if k not in name_to_idx:
            # 忽略不存在的列，后续写入留空
            pass

    # 额外读取用于总分/建议集数计算的分项
    extras = [
        '剧情结构（得分）', '角色塑造（得分）', '冲突设置（得分）', '台词质量（得分）',
        '人物饱满度（得分）', '制作（得分）', '宣发（得分）',
        '商业价值｜题材热度', '商业价值｜传播因子', '商业价值｜可生产性', '商业价值｜合规风险'
    ]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or (r[0] is None and r[name_to_idx.get('剧本名', 1)] is None):
            continue
        row = {k: (r[name_to_idx[k]] if k in name_to_idx else None) for k in required}
        for k in extras:
            row[k] = r[name_to_idx[k]] if k in name_to_idx else None
        rows.append(row)
    return rows


def to_float(v):
    try:
        return float(str(v).replace('%', '').strip())
    except Exception:
        return 0.0


def recommend_platform(倾向: str, 受众定位: str, 营销方向: str, 上线时间: str) -> str:
    t = (倾向 or '').strip()
    aud = (受众定位 or '').strip()
    mk = (营销方向 or '').strip()
    win = (上线时间 or '').strip()

    rec = []
    if '45+' in aud or '下沉' in aud:
        rec += ['视频号', '快手']
    if '18–35' in aud or '一二线' in aud:
        rec += ['抖音', '快手']
    if '女频' in t:
        rec = ['抖音', '小红书'] + ([rec[0]] if rec else [])
    if '甜虐' in mk or '甜宠' in mk:
        rec = ['抖音', '小红书']
    if '反转爽点' in mk or '成长救赎' in mk:
        rec = ['抖音', '快手']
    if '工作日晚间' in win:
        rec = (rec or ['视频号', '抖音'])
    if '周末黄金' in win:
        rec = (rec or ['抖音', '快手'])
    if '节假日前后' in win:
        rec = (rec or ['抖音', '快手', '视频号'])

    # 去重保持顺序
    seen = set()
    ordered = []
    for x in rec:
        if x not in seen:
            seen.add(x)
            ordered.append(x)
    return '/'.join(ordered) if ordered else '抖音/快手'


def build_remark(mk: str, win: str, aud: str, biz: Any, rec: str) -> str:
    return f"主卖点：{mk or ''}；窗口：{win or ''}；人群：{aud or ''}；平台：{rec}；商业指数：{biz or ''}"


def write_ranking(ws, rows: List[Dict[str, Any]]):
    ws.delete_rows(1, ws.max_row)
    header = [
        '排名', '序号', '剧本名', '总分', '评分等级', '商业价值指数', '倾向', '时代背景', '主题',
        '字数（千字）', '建议集数', '平台推荐', '营销方向', '上线时间', '受众定位', '备注'
    ]
    ws.append(header)
    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row.get('序号'),
            row.get('剧本名'),
            row.get('总分'),
            row.get('评分等级'),
            row.get('商业价值指数'),
            row.get('倾向'),
            row.get('时代背景'),
            row.get('主题'),
            row.get('字数（千字）'),
            row.get('建议集数'),
            row.get('平台推荐'),
            row.get('营销方向'),
            row.get('上线时间'),
            row.get('受众定位'),
            row.get('备注'),
        ])
    ws.freeze_panes = 'A2'
    widths = [6, 6, 28, 8, 10, 10, 8, 10, 16, 10, 10, 14, 14, 12, 18, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _calc_suggest_episodes(words_k: Optional[Any]) -> str:
    w = to_float(words_k)
    if w <= 8:
        return '15–20集'
    if w <= 20:
        return '20–30集'
    return '30–40集'


def _calc_biz_index(ax: Any, ay: Any, az: Any, ba: Any) -> float:
    return round((to_float(ax)/2*0.4 + to_float(ay)/1.5*0.3 + to_float(az)/1*0.2 + to_float(ba)/0.5*0.1)*100, 2)


def _calc_total(ac: Any, ah: Any, ai: Any, al: Any, aq: Any, aS: Any, aw: Any, biz_index: float) -> float:
    return round(to_float(ac) + to_float(ah) + to_float(ai) + to_float(al) + to_float(aq) + to_float(aS) + to_float(aw) + (biz_index/100*5), 2)


def _grade_by_score(total: float) -> str:
    if total <= 20:
        return '极弱'
    if total <= 40:
        return '偏弱'
    if total <= 60:
        return '中等'
    if total <= 80:
        return '较强'
    return '卓越'


def main(path: str = SRC_PATH, src_sheet: str = SRC_SHEET, out_sheet: str = RANK_SHEET, top_n: int = 10):
    # 使用 data_only=True 读取公式计算后的值，确保“总分/商业价值指数”等为数值
    wb = load_workbook(path, data_only=True)
    src = wb[src_sheet]
    rows = read_sheet(src)
    # 补算建议集数、商业价值指数与总分、评分等级（避免公式未计算导致空值）
    for r in rows:
        # 建议集数
        if not r.get('建议集数'):
            r['建议集数'] = _calc_suggest_episodes(r.get('字数（千字）'))
        # 商业价值指数
        biz = r.get('商业价值指数')
        if biz in (None, ''):
            biz = _calc_biz_index(r.get('商业价值｜题材热度'), r.get('商业价值｜传播因子'), r.get('商业价值｜可生产性'), r.get('商业价值｜合规风险'))
            r['商业价值指数'] = biz
        else:
            biz = to_float(biz)
        # 总分
        total = r.get('总分')
        if total in (None, '') or to_float(total) == 0.0:
            total = _calc_total(
                r.get('剧情结构（得分）'), r.get('角色塑造（得分）'), r.get('冲突设置（得分）'), r.get('台词质量（得分）'),
                r.get('人物饱满度（得分）'), r.get('制作（得分）'), r.get('宣发（得分）'), biz
            )
            r['总分'] = total
        else:
            total = to_float(total)
        # 评分等级
        if not r.get('评分等级'):
            r['评分等级'] = _grade_by_score(total)
    # 排序
    rows.sort(key=lambda r: to_float(r.get('总分')), reverse=True)
    if top_n:
        rows = rows[:top_n]
    # 填充平台推荐与备注
    for r in rows:
        rec = recommend_platform(r.get('倾向'), r.get('受众定位'), r.get('营销方向'), r.get('上线时间'))
        if not r.get('平台推荐'):
            r['平台推荐'] = rec
        r['备注'] = build_remark(r.get('营销方向'), r.get('上线时间'), r.get('受众定位'), r.get('商业价值指数'), r['平台推荐'])
    # 写入
    ws = wb[out_sheet] if out_sheet in wb.sheetnames else wb.create_sheet(out_sheet)
    write_ranking(ws, rows)
    wb.save(path)
    print(f"[完成] 已生成排序视图：{out_sheet}（Top {len(rows)}）")


if __name__ == '__main__':
    args = sys.argv[1:]
    kwargs = {}
    if args:
        kwargs['path'] = args[0]
    if len(args) > 1:
        kwargs['src_sheet'] = args[1]
    if len(args) > 2:
        kwargs['out_sheet'] = args[2]
    if len(args) > 3:
        kwargs['top_n'] = int(args[3])
    main(**kwargs)