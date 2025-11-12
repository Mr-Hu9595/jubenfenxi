#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
修复指定 Excel 文件的D列“剧本概要”，并生成带 v2 标识的新文件：

 - 保留原始文件备份：original_path + .backup_YYYYMMDD_HHMMSS.xlsx
 - 在同目录生成 *_v2.xlsx 文件，D列改为约500字摘要（含质量校验）。
 - 其他列原样复制；为稳健性统一回填公式联动。

用法：
  python3 scripts/repair_excel_v2.py /Users/mr.hu/Desktop/剧本分析/bug/test1_20251112_014616_剧本分析.xlsx
"""

import os
import sys
import time
from openpyxl import load_workbook, Workbook

BASE = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(BASE, 'tools'))

from summary_quality import generate_summary
import universal_cli as uni


def main():
    if len(sys.argv) < 2:
        print("[用法] python3 scripts/repair_excel_v2.py <excel_path>")
        sys.exit(1)
    src = sys.argv[1]
    if not os.path.exists(src):
        print(f"[错误] 文件不存在：{src}")
        sys.exit(2)

    dirname = os.path.dirname(src)
    name, ext = os.path.splitext(os.path.basename(src))
    ts = time.strftime('%Y%m%d_%H%M%S')
    backup = os.path.join(dirname, f"{name}.backup_{ts}{ext}")
    dst = os.path.join(dirname, f"{name}_v2{ext}")

    # 备份原始文件
    try:
        import shutil
        shutil.copy2(src, backup)
        print(f"[备份] 已保存：{backup}")
    except Exception as e:
        print(f"[警告] 备份失败：{e}")

    wb = load_workbook(src, data_only=False)
    sheet = '工作表1'
    if sheet not in wb.sheetnames:
        print("[错误] 未找到工作表：工作表1")
        sys.exit(3)
    ws = wb[sheet]

    # 创建新工作簿并复制表头
    nwb = Workbook()
    nws = nwb.active
    nws.title = sheet
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        val = ws.cell(row=1, column=c).value
        if c == 4:
            val = "剧本概要"  # 统一表头
        nws.cell(row=1, column=c, value=val)

    # 逐行复制，D列改为摘要
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        title = ws.cell(row=r, column=2).value or ""
        text = ws.cell(row=r, column=4).value or ""
        for c in range(1, max_col + 1):
            if c == 4:
                try:
                    summary = generate_summary(str(text), str(title), target_chars=500)
                except Exception:
                    summary = (str(text)[:500]).strip()
                nws.cell(row=r, column=c, value=summary)
            else:
                nws.cell(row=r, column=c, value=ws.cell(row=r, column=c).value)

    # 回填公式联动
    try:
        uni.apply_formulas(nws, 2, nws.max_row)
    except Exception:
        pass

    nwb.save(dst)
    print(f"[完成] 已生成修复文件：{dst}")


if __name__ == '__main__':
    main()