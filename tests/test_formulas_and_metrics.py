import os
import sys

HERE = os.path.dirname(__file__)
ROOT = os.path.abspath(os.path.join(HERE, '..'))
TOOLS = os.path.join(ROOT, 'tools')
if TOOLS not in sys.path:
    sys.path.insert(0, TOOLS)

from openpyxl import Workbook

import universal_cli as uni
from auto_score_from_text import character_fullness_components


def test_apply_formulas_persona_fullness_and_total_score():
    wb = Workbook()
    ws = wb.active
    ws.title = '评估输入'
    # 准备一行：设置人物饱满度四组件，制作与宣发与商业值
    r = 2
    ws[f'AM{r}'] = 80  # 关键词权重
    ws[f'AN{r}'] = 50  # 连接词权重
    ws[f'AO{r}'] = 30  # 反差词对
    ws[f'AP{r}'] = 10  # 事件层级与弧光
    # 内容四项分与制作/宣发，用于总分公式拼装
    ws[f'AC{r}'] = 12
    ws[f'AH{r}'] = 13
    ws[f'AI{r}'] = 9
    ws[f'AL{r}'] = 8
    ws[f'AS{r}'] = 14
    ws[f'AW{r}'] = 9
    ws[f'AX{r}'] = 2
    ws[f'AY{r}'] = 1.5
    ws[f'AZ{r}'] = 1
    ws[f'BA{r}'] = 0.5

    uni.apply_formulas(ws, r, r)

    # AQ 应为 0–100 指数的加权公式
    assert ws[f'AQ{r}'].value == (
        f'=ROUND(AM{r}*0.4 + AN{r}*0.3 + AO{r}*0.2 + AP{r}*0.1,2)'
    )
    # AR 标签公式基于 AQ（0–100）
    assert ws[f'AR{r}'].value == (
        f'=IF(AQ{r}>=67,"高",IF(AQ{r}>=34,"中","低"))'
    )
    # BB 的加权归一公式保持不变
    assert ws[f'BB{r}'].value == (
        f'=ROUND((AX{r}/2*0.4 + AY{r}/1.5*0.3 + AZ{r}/1*0.2 + BA{r}/0.5*0.1)*100,2)'
    )
    # BC 应包含 AQ/100*20 的折算
    assert ws[f'BC{r}'].value == (
        f'=ROUND(AC{r}+AH{r}+AI{r}+AL{r}+(AQ{r}/100*20)+AS{r}+AW{r}+(BB{r}/100*5),2)'
    )


def test_character_fullness_components_basic():
    text = '反转 成长 但是 然而 看似 实则 温柔 暴烈 救赎 黑化 洗白 事件 转变 觉醒'
    comp = character_fullness_components(text)
    # 四组件应存在且为 0–100 范围
    keys = [
        '人物饱满度｜关键词权重',
        '人物饱满度｜连接词权重',
        '人物饱满度｜反差词对',
        '人物饱满度｜事件层级与弧光',
    ]
    for k in keys:
        assert k in comp
        assert 0 <= comp[k] <= 100